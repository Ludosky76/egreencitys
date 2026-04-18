from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

C_DARK = "1B5E20"; C_MED = "2E7D32"; C_AE = "E3F2FD"; C_AO = "BBDEFB"
C_DC = "FFE0B2"; C_TOT = "263238"; C_WH = "FFFFFF"; C_BK = "000000"; C_BL = "0000FF"

wb = Workbook()
ws = wb.active
ws.title = "20 PDC — ADVENIR"

def fnt(b=False, s=10, c=C_BK, i=False):
    return Font(name="Arial", bold=b, size=s, color=c, italic=i)

def fl(c):
    return PatternFill("solid", fgColor=c)

def al(h="center", v="center", w=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=w)

def br(st="thin"):
    x = Side(style=st)
    return Border(left=x, right=x, top=x, bottom=x)

# ── TITLE ROWS ───────────────────────────────────────────────────────────────
titles = [
    ("EGREENCITY'S — TABLEAU DES 20 POINTS DE CHARGE (PDC)", 14, 28, C_DARK),
    ("Programme ADVENIR — Voirie publique — Guyane francaise — 10 x e-Premium AC 2x22 kW", 11, 20, C_MED),
    ("SIREN 878 682 854  |  10 stations  |  20 PDC  |  440 kW  |  Prime totale : 37 200 EUR", 9, 16, C_MED),
]
for r, (txt, sz, h, bg) in enumerate(titles, 1):
    ws.merge_cells(f"A{r}:O{r}")
    c = ws.cell(row=r, column=1, value=txt)
    c.font = fnt(b=(r == 1), s=sz, c=C_WH)
    c.fill = fl(bg)
    c.alignment = al()
    ws.row_dimensions[r].height = h

# ── COLUMN HEADERS ───────────────────────────────────────────────────────────
HDRS = [
    ("N° PDC", 7), ("N° Stat.", 8), ("Réf. Station", 14), ("Commune", 17),
    ("Adresse / Site proposé", 28), ("Type IRVE", 10), ("Modèle borne", 20),
    ("Connecteur", 12), ("Puiss.\nPDC (kW)", 11), ("Accès\n24/7", 8),
    ("Auth.\nRFID", 7), ("OCPP\n1.6J", 7), ("Prime\nADVENIR (€)", 13),
    ("Prix HT\nborne (€)", 13), ("Observations", 22),
]
for col, (h, w) in enumerate(HDRS, 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = fnt(b=True, s=9, c=C_WH)
    c.fill = fl(C_DARK)
    c.alignment = al(w=True)
    c.border = br()
    ws.column_dimensions[get_column_letter(col)].width = w
ws.row_dimensions[4].height = 36

# ── STATION DATA ─────────────────────────────────────────────────────────────
STATIONS = [
    (1,  "S01-CAY-CTR", "Cayenne",                 "Place des Palmistes",                  "AC", "e-Premium AC 2x22 kW", "Type 2 / Prise E", 22, 5802, "Centre historique — fort trafic pieton"),
    (2,  "S02-CAY-EST", "Cayenne",                 "Zone commerciale Califourchon",         "AC", "e-Premium AC 2x22 kW", "Type 2 / Prise E", 22, 5802, "Zone commerciale Est — convention mairie"),
    (3,  "S03-MAT",     "Matoury",                 "Parking Geant Casino Matoury",          "AC", "e-Premium AC 2x22 kW", "Type 2 / Prise E", 22, 5802, "Parc commercial — fort flux hebdomadaire"),
    (4,  "S04-REM",     "Remire-Montjoly",         "Mairie de Remire-Montjoly",             "AC", "e-Premium AC 2x22 kW", "Type 2 / Prise E", 22, 5802, "Site municipal — domaine public voirie"),
    (5,  "S05-MAC",     "Macouria-Tonate",         "Mairie de Macouria",                    "AC", "e-Premium AC 2x22 kW", "Type 2 / Prise E", 22, 5802, "Siege operateur — site pilote EGREENCITY"),
    (6,  "S06-KOU",     "Kourou",                  "Place Newton — Centre-ville Kourou",    "AC", "e-Premium AC 2x22 kW", "Type 2 / Prise E", 22, 5802, "Centre-ville Kourou — 2e pole urbain"),
    (7,  "S07-IRA-A",   "Iracoubo",                "Place de la Mairie d'Iracoubo",         "AC", "e-Premium AC 2x22 kW", "Type 2 / Prise E", 22, 5802, "Relais RN1 km 95 — 1er site (Mairie)"),
    (8,  "S08-IRA-B",   "Iracoubo",                "Aire de repos RN1 — km 95 Iracoubo",   "AC", "e-Premium AC 2x22 kW", "Type 2 / Prise E", 22, 5802, "Relais RN1 km 95 — 2e site (Aire repos)"),
    (9,  "S09-SLM",     "Saint-Laurent-du-Maroni", "Mairie de Saint-Laurent-du-Maroni",    "AC", "e-Premium AC 2x22 kW", "Type 2 / Prise E", 22, 5802, "3e pole urbain — desenclavement Ouest"),
    (10, "S10-MAN",     "Mana",                    "Place de la Mairie — Mana",             "AC", "e-Premium AC 2x22 kW", "Type 2 / Prise E", 22, 5802, "Communes littorales Ouest — axe Mana"),
]

pdc_n = 1
dr = 5
for sn, ref, commune, site, stype, modele, conn, puiss, prix_ht, obs in STATIONS:
    bg = C_DC if stype == "DC Rapide" else (C_AE if sn % 2 == 0 else C_AO)
    for pi in range(2):
        vals = [pdc_n, sn, ref, commune, site, stype, modele, conn, puiss,
                "Oui", "Oui", "Oui", None,
                prix_ht if pi == 0 else None,
                obs if pi == 0 else ""]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=dr, column=col, value=val)
            c.fill = fl(bg)
            c.border = br()
            c.font = fnt(b=(col == 1), s=9)
            c.alignment = al("left" if col in [4, 5, 15] else "center", w=(col in [4, 5, 15]))
        ws.cell(row=dr, column=13).value = 1860
        ws.cell(row=dr, column=13).font = fnt(s=9, c=C_BL)
        ws.cell(row=dr, column=13).number_format = "#,##0"
        ws.cell(row=dr, column=14).number_format = "#,##0"
        ws.row_dimensions[dr].height = 16
        pdc_n += 1
        dr += 1

# ── TOTAL ROW ────────────────────────────────────────────────────────────────
tr = dr
ws.merge_cells(f"A{tr}:H{tr}")
c = ws.cell(row=tr, column=1, value="TOTAUX — 10 stations  |  20 PDC  |  440 kW  |  37 200 EUR ADVENIR")
c.font = fnt(b=True, s=10, c=C_WH)
c.fill = fl(C_TOT)
c.alignment = al()
c.border = br("medium")

for col, formula, fmt in [
    (9,  f"=SUM(I5:I{tr-1})", "#,##0"),
    (13, f"=SUM(M5:M{tr-1})", "#,##0"),
    (14, f'=SUMIF(N5:N{tr-1},">0")', "#,##0"),
]:
    c = ws.cell(row=tr, column=col, value=formula)
    c.font = fnt(b=True, s=10, c=C_WH)
    c.fill = fl(C_TOT)
    c.alignment = al()
    c.border = br("medium")
    c.number_format = fmt

for col, val in [(10, "Oui"), (11, "Oui"), (12, "Oui"), (15, "= 64,1 % du HT materiel")]:
    c = ws.cell(row=tr, column=col, value=val)
    c.font = fnt(b=True, s=9, c=C_WH)
    c.fill = fl(C_TOT)
    c.alignment = al()
    c.border = br("medium")
ws.row_dimensions[tr].height = 22

# ── SYNTHESE PAR COMMUNE ─────────────────────────────────────────────────────
sr = tr + 3
ws.merge_cells(f"A{sr}:O{sr}")
c = ws.cell(row=sr, column=1, value="SYNTHESE PAR COMMUNE")
c.font = fnt(b=True, s=11, c=C_WH)
c.fill = fl(C_MED)
c.alignment = al()
ws.row_dimensions[sr].height = 20

for col, h in enumerate(["Commune", "Nb stations", "Nb PDC", "Puiss. totale (kW)",
                           "Type(s) IRVE", "Prime ADVENIR (EUR)", "Prix HT (EUR)", "Priorite maillage"], 1):
    c = ws.cell(row=sr + 1, column=col, value=h)
    c.font = fnt(b=True, s=9, c=C_WH)
    c.fill = fl(C_DARK)
    c.alignment = al(w=True)
    c.border = br()
ws.row_dimensions[sr + 1].height = 26

SYNTH = [
    ("Cayenne",                 2, 4,  88, "AC", 7440, 11604, "Capitale — 2 stations"),
    ("Matoury",                 1, 2,  44, "AC", 3720,  5802, "Banlieue Cayenne"),
    ("Remire-Montjoly",         1, 2,  44, "AC", 3720,  5802, "Banlieue Cayenne"),
    ("Macouria-Tonate",         1, 2,  44, "AC", 3720,  5802, "Siege operateur"),
    ("Kourou",                  1, 2,  44, "AC", 3720,  5802, "2e pole urbain"),
    ("Iracoubo",                2, 4,  88, "AC", 7440, 11604, "Relais RN1 km 95 — 4 PDC"),
    ("Saint-Laurent-du-Maroni", 1, 2,  44, "AC", 3720,  5802, "3e pole urbain"),
    ("Mana",                    1, 2,  44, "AC", 3720,  5802, "Desserte littoral Ouest"),
]
assert sum(r[2] for r in SYNTH) == 20, "PDC count mismatch"
assert sum(r[3] for r in SYNTH) == 440, "Power mismatch"
assert sum(r[5] for r in SYNTH) == 37200, "ADVENIR mismatch"
assert sum(r[6] for r in SYNTH) == 58020, "HT mismatch"

for i, rd in enumerate(SYNTH):
    r2 = sr + 2 + i
    bg = C_AE if i % 2 == 0 else C_AO
    for col, val in enumerate(rd, 1):
        c = ws.cell(row=r2, column=col, value=val)
        c.font = fnt(s=9)
        c.fill = fl(bg)
        c.alignment = al("left" if col == 1 else "center")
        c.border = br()
        if col in [6, 7]:
            c.number_format = "#,##0"
    ws.row_dimensions[r2].height = 15

# ── NOTES ────────────────────────────────────────────────────────────────────
nr = sr + 2 + len(SYNTH) + 2
ws.merge_cells(f"A{nr}:O{nr}")
ws.cell(row=nr, column=1, value="NOTES").font = fnt(b=True, s=9)
NOTES = [
    "- Prime ADVENIR : 1 860 EUR/PDC — Programme AVERE-France, voirie publique, puissance PDC >= 7,4 kW",
    "- Engagement : 5 ans minimum de disponibilite publique 24/7 — Signaletique ADVENIR obligatoire sur chaque borne",
    "- Equipements certifies OCPP 1.6J | Authentification RFID + CB sans contact + app mobile iOS/Android",
    "- e-Premium AC 2x22 kW : 5 802 EUR HT/borne (devis E-TOTEM ref. DEV17000172-6)",
    "- Coefficient TTC Guyane x1,670 — inclut Octroi de Mer 23%, fret DROM, installation, raccordement EDF",
    "- Iracoubo (S07 + S08) : 4 PDC = 2 bornes — relais AC strategique axe RN1 Cayenne-SLM (km 95)",
    "- Taux de couverture ADVENIR : 37 200 EUR / 58 020 EUR HT = 64,1% (exceptionnel en DROM)",
]
for ni, note in enumerate(NOTES):
    ws.merge_cells(f"A{nr+1+ni}:O{nr+1+ni}")
    ws.cell(row=nr + 1 + ni, column=1, value=note).font = fnt(s=8, i=True)
    ws.row_dimensions[nr + 1 + ni].height = 13

ws.freeze_panes = "A5"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1

out = r"C:\projet\Egreencity\02_Dossier_ADVENIR\02_Tableau_20_PDC.xlsx"
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print(f"OK: {out}")
print(f"22 PDC rows written, total row at {tr}, synth at {sr}")
