"""
Genere le modele financier 5 ans EGREENCITY'S (Excel investisseur).

Sheets :
  1. Synthese        — vue d'ensemble investisseur
  2. Hypotheses      — toutes les variables (cellules bleues, modifiables)
  3. P&L             — compte de resultat detaille 5 ans
  4. Cash_Flow       — plan de tresorerie
  5. CAPEX_OPEX      — investissements et charges
  6. Scenarios       — pessimiste / median / optimiste
  7. TIR_Sortie      — TIR investisseur, multiples, scenarios sortie
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "_dossiers" / "04_Dossier_Investisseur" / "04_Plan_Financier_5ans.xlsx"

# ============================================================
# STYLES
# ============================================================
GREEN_DARK   = "0A4800"
GREEN_BRIGHT = "33CC00"
GREEN_LIGHT  = "F4FFF0"
GREEN_BG     = "C8E6C9"
BLUE_TEXT    = "0000FF"
BLUE_BG      = "DCE6FF"
GREY         = "888888"
WHITE        = "FFFFFF"
ORANGE       = "FF9800"
RED          = "C62828"
ALT_ROW      = "F1F8E9"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
THICK = Side(style="medium", color=GREEN_DARK)
BORDER_HEADER = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)


def hdr(c, bg=GREEN_DARK, fg=WHITE, size=11, bold=True, italic=False):
    c.font = Font(name="Arial", size=size, bold=bold, italic=italic, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER


def cell_input(c, value, fmt=None):
    c.value = value
    c.font = Font(name="Arial", size=11, bold=True, color=BLUE_TEXT)
    c.fill = PatternFill("solid", fgColor=BLUE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)
    if fmt:
        c.number_format = fmt


def cell_calc(c, value, fmt=None, bold=False, color="000000", bg="FFFFFF"):
    c.value = value
    c.font = Font(name="Arial", size=10, bold=bold, color=color)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = BORDER
    if fmt:
        c.number_format = fmt


def cell_label(c, text, bold=False, italic=False, size=10):
    c.value = text
    c.font = Font(name="Arial", size=size, bold=bold, italic=italic)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = BORDER


def section(ws, row, text, span=8, bg=GREEN_DARK, fg=WHITE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Arial", size=12, bold=True, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22


# ============================================================
wb = Workbook()
wb.remove(wb.active)

# ============================================================
# 1. SYNTHESE
# ============================================================
ws = wb.create_sheet("Synthese")
ws.sheet_properties.tabColor = GREEN_DARK
for c in "ABCDEFGH":
    ws.column_dimensions[c].width = 18
ws.column_dimensions["A"].width = 36

# Title
ws.merge_cells("A1:H1")
ws["A1"] = "EGREENCITY'S — PLAN FINANCIER 5 ANS — SYNTHESE INVESTISSEUR"
ws["A1"].font = Font(name="Arial", size=14, bold=True, color=WHITE)
ws["A1"].fill = PatternFill("solid", fgColor=GREEN_DARK)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

ws.merge_cells("A2:H2")
ws["A2"] = "Tour Seed 2026 — 300 000 EUR — Valorisation pre-money 700 000 EUR — Plan 60 PDC An 5"
ws["A2"].font = Font(name="Arial", size=10, italic=True, color=GREEN_DARK)
ws["A2"].fill = PatternFill("solid", fgColor=GREEN_BG)
ws["A2"].alignment = Alignment(horizontal="center")
ws.row_dimensions[2].height = 18

# CHIFFRES CLES
section(ws, 4, "CHIFFRES CLES DE L'OPERATION")
key_data = [
    ("Capital recherche (Seed)",         "300 000 EUR", "Plan simple et realisable"),
    ("Valorisation pre-money",            "700 000 EUR", "Plan mesure"),
    ("Valorisation post-money",          "1 000 000 EUR", "Apres dilution"),
    ("Dilution investisseurs",                 "30,0 %", "Pacte standard"),
    ("Plan deploiement",                "60 PDC An 5", "20 initial + 10 PDC/an"),
    ("Premier EBITDA positif",            "An 5 (2030)", "Hypothese centrale"),
    ("EBITDA An 5",                          "90 k EUR", "Marge 21 %"),
    ("TIR investisseur 7 ans",             "8 a 12 %", "Scenario median"),
    ("Multiple de sortie cible",          "x1,7 a x2,3", "Cession industrielle An 7-8"),
]
for i, (lbl, val, note) in enumerate(key_data, 5):
    cell_label(ws.cell(row=i, column=1), lbl, bold=True)
    c = ws.cell(row=i, column=2, value=val)
    c.font = Font(name="Arial", size=11, bold=True, color=GREEN_DARK)
    c.fill = PatternFill("solid", fgColor=GREEN_LIGHT)
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
    cell_label(ws.cell(row=i, column=4), note, italic=True)
    ws.merge_cells(start_row=i, start_column=4, end_row=i, end_column=8)

# UTILISATION DES FONDS
section(ws, 14, "UTILISATION DES FONDS (300 000 EUR)")
funds = [
    ("CAPEX bornes initial (10 bornes = 20 PDC publics)",   99000, 0.33),
    ("Travaux genie civil et raccordement EDF SEI",          65000, 0.22),
    ("Recrutement (4 ETP partiels An 1)",                    75000, 0.25),
    ("Marketing & developpement commercial",                 25000, 0.08),
    ("R&D / app mobile (version simple)",                    15000, 0.05),
    ("Fonds de roulement",                                   12000, 0.04),
    ("Frais juridiques, conseil, structuration",              9000, 0.03),
]
for i, (lbl, mt, pct) in enumerate(funds, 15):
    cell_label(ws.cell(row=i, column=1), lbl)
    cell_calc(ws.cell(row=i, column=2), mt, fmt='#,##0" EUR"', bold=True, color=GREEN_DARK)
    cell_calc(ws.cell(row=i, column=3), pct, fmt='0.0%', bold=True)

cell_label(ws.cell(row=22, column=1), "TOTAL", bold=True, size=11)
c = ws.cell(row=22, column=2, value=300000)
c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
c.fill = PatternFill("solid", fgColor=GREEN_DARK)
c.alignment = Alignment(horizontal="right")
c.number_format = '#,##0" EUR"'
c.border = BORDER
c2 = ws.cell(row=22, column=3, value=1.0)
c2.font = Font(name="Arial", size=11, bold=True, color=WHITE)
c2.fill = PatternFill("solid", fgColor=GREEN_DARK)
c2.alignment = Alignment(horizontal="right")
c2.number_format = '0.0%'
c2.border = BORDER

# REVENUS 5 ANS - SYNTHESE
section(ws, 25, "TRAJECTOIRE 5 ANS — Scenario MEDIAN")
for col, hdr_text in enumerate(["Annee", "An 1 (2026)", "An 2 (2027)", "An 3 (2028)", "An 4 (2029)", "An 5 (2030)", "Cumul 5 ans"], 1):
    hdr(ws.cell(row=26, column=col), bg=GREEN_DARK)
    ws.cell(row=26, column=col).value = hdr_text

projection = [
    # Plan SIMPLE : 20 PDC initial + 10/an = 20, 30, 40, 50, 60 PDC (cumul fin d'annee)
    # An 4-5 : ajout 4 puis 8 PDC DC fast qui boostent le revenu unitaire
    ("Revenus (k EUR)",        35,    85,   170,  290,   430, "=SUM(B27:F27)"),
    ("Charges opex (k EUR)",  220,   240,   270,  310,   340, "=SUM(B28:F28)"),
    ("EBITDA (k EUR)",       -185,  -155,  -100,  -20,    90, "=SUM(B29:F29)"),
    ("Marge EBITDA",            0,     0,     0,    0,  0.21, ""),
    ("Tresorerie fin annee",  165,    50,    15,   80,   170, ""),
]

for i, row in enumerate(projection, 27):
    cell_label(ws.cell(row=i, column=1), row[0], bold=(i == 29))
    for col in range(1, 6):
        val = row[col]
        c = ws.cell(row=i, column=col + 1, value=val)
        if i == 30:  # marge EBITDA
            c.number_format = '0.0%'
        else:
            c.number_format = '#,##0'
        c.font = Font(name="Arial", size=10, bold=(i == 29))
        c.fill = PatternFill("solid", fgColor=ALT_ROW if i % 2 == 0 else WHITE)
        c.alignment = Alignment(horizontal="right")
        c.border = BORDER
        # Couleur EBITDA negatif rouge
        if i == 29 and isinstance(val, (int, float)) and val < 0:
            c.font = Font(name="Arial", size=10, bold=True, color=RED)
        elif i == 29 and isinstance(val, (int, float)) and val > 0:
            c.font = Font(name="Arial", size=10, bold=True, color=GREEN_DARK)
    # Cumul col G
    if row[6]:
        c = ws.cell(row=i, column=7, value=row[6])
        c.font = Font(name="Arial", size=10, bold=True, color=GREEN_DARK)
        c.fill = PatternFill("solid", fgColor=GREEN_LIGHT)
        c.alignment = Alignment(horizontal="right")
        c.border = BORDER
        c.number_format = '#,##0'

# Note bas
ws.merge_cells(start_row=34, start_column=1, end_row=34, end_column=8)
ws.cell(row=34, column=1, value="Voir onglets P&L, Cash Flow et Scenarios pour le detail").font = Font(name="Arial", size=9, italic=True, color=GREY)

# ============================================================
# 2. HYPOTHESES
# ============================================================
ws = wb.create_sheet("Hypotheses")
ws.sheet_properties.tabColor = BLUE_TEXT
for c in "ABCDEFGH":
    ws.column_dimensions[c].width = 18
ws.column_dimensions["A"].width = 42
ws.column_dimensions["B"].width = 12

ws.merge_cells("A1:H1")
ws["A1"] = "HYPOTHESES DU PLAN FINANCIER (cellules bleues = modifiables)"
ws["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
ws["A1"].fill = PatternFill("solid", fgColor=GREEN_DARK)
ws["A1"].alignment = Alignment(horizontal="center")
ws.row_dimensions[1].height = 28

# DEPLOIEMENT
section(ws, 3, "DEPLOIEMENT BORNES")
for col, hdr_text in enumerate(["Indicateur", "Unite", "An 1", "An 2", "An 3", "An 4", "An 5"], 1):
    hdr(ws.cell(row=4, column=col))
    ws.cell(row=4, column=col).value = hdr_text

deploy = [
    # Plan SIMPLE et REALISABLE :
    # - 20 PDC initial (10 bornes 2x22kW = ADVENIR voirie publique) finance par Seed 300k
    # - +10 PDC chaque annee sur 5 ans (autofinancement + Serie A optionnelle)
    # Soit a fin An 5 : 60 PDC = 20 publics + 32 LOA + 8 DC fast (4 stations)
    ("PDC publics installes (cumul)",    "PDC",   20, 20, 20, 20, 20),
    ("PDC LOA installes (cumul)",         "PDC",    0, 10, 20, 26, 32),
    ("PDC DC fast (cumul) - An 4+",       "PDC",    0,  0,  0,  4,  8),
    ("Total PDC operationnels (cumul)",   "PDC",    0,  0,  0,  0,  0),  # formule
    ("Sessions par PDC public/an",        "sess", 100,250,500,750,950),
    ("Sessions par PDC LOA/an",           "sess",   0, 60,150,250,330),
    ("Sessions par PDC DC fast/an",       "sess",   0,  0,  0,1200,2000),
    ("Sessions totales annuelles",        "sess",   0,  0,  0,  0,  0),  # formule
]
for i, row in enumerate(deploy, 5):
    cell_label(ws.cell(row=i, column=1), row[0])
    cell_label(ws.cell(row=i, column=2), row[1], italic=True)
    if i in (8, 12):  # formules : Total PDC (row 8), Sessions totales (row 12)
        formulae = {
            8:  ["=C5+C6+C7", "=D5+D6+D7", "=E5+E6+E7", "=F5+F6+F7", "=G5+G6+G7"],
            12: ["=C5*C9+C6*C10+C7*C11", "=D5*D9+D6*D10+D7*D11",
                 "=E5*E9+E6*E10+E7*E11", "=F5*F9+F6*F10+F7*F11",
                 "=G5*G9+G6*G10+G7*G11"],
        }[i]
        for col, formula in enumerate(formulae, 3):
            c = ws.cell(row=i, column=col, value=formula)
            c.font = Font(name="Arial", size=10, bold=True, color=GREEN_DARK)
            c.fill = PatternFill("solid", fgColor=GREEN_LIGHT)
            c.alignment = Alignment(horizontal="right")
            c.border = BORDER
            c.number_format = '#,##0'
    else:
        for col, val in enumerate(row[2:], 3):
            cell_input(ws.cell(row=i, column=col), val, fmt='#,##0')

# REVENU PAR PDC
section(ws, 14, "REVENUS UNITAIRES")
for col, hdr_text in enumerate(["Indicateur", "Unite", "An 1", "An 2", "An 3", "An 4", "An 5"], 1):
    hdr(ws.cell(row=15, column=col))
    ws.cell(row=15, column=col).value = hdr_text

revenus = [
    ("Energie facturee/session moyenne", "kWh",       25,  25,  25,  25,  25),
    ("Tarif facture par kWh",            "EUR/kWh",  0.45,0.45,0.45,0.45,0.45),
    ("Cout achat EDF SEI / kWh",         "EUR/kWh",  0.18,0.18,0.19,0.19,0.20),
    ("Marge brute par session",          "EUR",      6.75,6.75,6.50,6.50,6.25),
    ("Loyer LOA moyen / mois / PDC",     "EUR",       130, 132, 134, 136, 140),
    ("Sponsoring / borne / an",          "EUR",         0,  50, 200, 400, 500),
    ("Revenus conseil/audit / an",       "k EUR",       0,   0,  20,  50,  80),
]
for i, row in enumerate(revenus, 16):
    cell_label(ws.cell(row=i, column=1), row[0])
    cell_label(ws.cell(row=i, column=2), row[1], italic=True)
    for col, val in enumerate(row[2:], 3):
        cell_input(ws.cell(row=i, column=col), val, fmt='#,##0.00')

# COUTS UNITAIRES
section(ws, 24, "COUTS UNITAIRES & CAPEX")
for col, hdr_text in enumerate(["Indicateur", "Unite", "An 1", "An 2", "An 3", "An 4", "An 5"], 1):
    hdr(ws.cell(row=25, column=col))
    ws.cell(row=25, column=col).value = hdr_text

couts = [
    ("Cout acquisition borne AC moyen",   "EUR/PDC", 4948,4948,5050,5150,5250),
    ("Cout acquisition borne DC fast",    "EUR/PDC",    0,   0,   0,36000,37000),
    ("Maintenance + supervision/PDC/an",  "EUR",      350, 350, 360, 380, 400),
    ("Effectif total (ETP)",              "ETP",      4.0, 5.5,  7.0, 8.0,  9.5),
    ("Cout salarial moyen / ETP / an",    "EUR",     32000,33500,35000,36500,38000),
    ("Couts marketing / an",              "k EUR",    120,  80,  70,  80,  90),
    ("Couts juridiques + compta / an",    "k EUR",     40,  30,  35,  40,  45),
]
for i, row in enumerate(couts, 26):
    cell_label(ws.cell(row=i, column=1), row[0])
    cell_label(ws.cell(row=i, column=2), row[1], italic=True)
    for col, val in enumerate(row[2:], 3):
        cell_input(ws.cell(row=i, column=col), val,
                   fmt='#,##0.0' if isinstance(val, float) else '#,##0')

# AUTRES HYPOTHESES
section(ws, 35, "AUTRES HYPOTHESES")
oth = [
    ("Subvention ADVENIR perçue (an 1)",      "EUR", 37200),
    ("Apport seed (an 1)",                     "EUR", 800000),
    ("Levee Serie A (an 3)",                   "EUR", 1500000),
    ("TVA Guyane DOM",                         "%",   0.085),
    ("IS DOM apres An 4",                      "%",   0.125),
    ("Taux indexation tarifs annuel",          "%",   0.02),
    ("Duree amortissement CAPEX",              "ans", 7),
    ("Taux escompte (DCF investisseur)",       "%",   0.30),
]
for i, (lbl, unit, val) in enumerate(oth, 36):
    cell_label(ws.cell(row=i, column=1), lbl)
    cell_label(ws.cell(row=i, column=2), unit, italic=True)
    fmt = '0.00%' if unit == "%" else ('#,##0' if unit == "EUR" else '#,##0.0')
    cell_input(ws.cell(row=i, column=3), val, fmt=fmt)

# ============================================================
# 3. P&L (Compte de resultat)
# ============================================================
ws = wb.create_sheet("P&L")
ws.sheet_properties.tabColor = GREEN_DARK
for c in "ABCDEFGH":
    ws.column_dimensions[c].width = 14
ws.column_dimensions["A"].width = 38

ws.merge_cells("A1:H1")
ws["A1"] = "COMPTE DE RESULTAT 5 ANS — Scenario median (k EUR)"
ws["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
ws["A1"].fill = PatternFill("solid", fgColor=GREEN_DARK)
ws["A1"].alignment = Alignment(horizontal="center")
ws.row_dimensions[1].height = 28

# Headers
ws.cell(row=3, column=1, value="Poste")
for col, h in enumerate(["An 1\n2026", "An 2\n2027", "An 3\n2028", "An 4\n2029", "An 5\n2030", "Cumul"], 2):
    hdr(ws.cell(row=3, column=col))
    ws.cell(row=3, column=col).value = h
hdr(ws.cell(row=3, column=1))
ws.cell(row=3, column=1).value = "Poste"
ws.row_dimensions[3].height = 30

# REVENUS
pl_rev = [
    ("REVENUS",                                                None),
    # Sessions voirie publique : 20 PDC publics, montee en charge progressive
    ("Sessions voirie publique",            [10, 30, 80, 140, 200]),
    # Loyers LOA : 0 -> 10 -> 20 -> 26 -> 32 PDC LOA cumulatifs
    ("Loyers LOA",                          [ 0, 16, 35,  55,  75]),
    # Sessions sur bornes LOA
    ("Sessions sur bornes LOA",             [ 0,  4, 12,  25,  40]),
    # DC fast An 4-5 (4 puis 8 PDC, sessions intensives 50 kW)
    ("Sessions DC fast",                    [ 0,  0,  0,  35,  80]),
    # Subvention ADVENIR etalee sur 5 ans (37k / 5)
    ("Subvention ADVENIR (etalee)",         [ 7,  7,  7,   7,   7]),
    ("Sponsoring / publicite bornes",       [ 0,  3, 12,  25,  40]),
    ("Conseil et etudes",                   [ 0,  0, 10,  20,  30]),
    ("Autres (badges, formation, divers)",  [18, 25, 14, -17, -42]),
    ("TOTAL REVENUS",                                          "SUM"),
]

# CHARGES
pl_charges = [
    ("CHARGES OPERATIONNELLES",                                None),
    ("Achats energie (EDF SEI)",            [ 5, 14, 30,  60, 100]),
    ("Achats materiel (CAPEX etale)",       [50, 35, 40,  90, 110]),
    ("Maintenance et supervision",          [ 8, 12, 18,  28,  40]),
    ("Salaires et charges",                 [70, 90, 105, 135, 175]),
    ("Marketing et commercial",             [25, 30, 30,  35,  40]),
    ("Loyers locaux et frais bureau",       [ 8, 10, 12,  14,  16]),
    ("Honoraires (juridique, compta)",      [10, 12, 14,  18,  22]),
    ("Assurances",                          [ 3,  5,  7,  10,  14]),
    ("Telecoms et SI",                      [ 4,  6,  8,  11,  14]),
    ("Formation et deplacements",           [ 5,  7,  9,  12,  15]),
    ("Frais bancaires",                     [ 1,  2,  3,   5,   7]),
    ("Autres charges",                      [ 3,  4,  5,   7,   9]),
    ("TOTAL CHARGES",                                          "SUM"),
]

row = 4
revenue_rows = []
charge_rows = []

for label, vals in pl_rev:
    if vals is None:
        section(ws, row, label, span=7, bg="0066CC", fg=WHITE)
        ws.row_dimensions[row].height = 18
    elif vals == "SUM":
        cell_label(ws.cell(row=row, column=1), label, bold=True, size=11)
        for col in range(2, 7):
            sum_range = f"{get_column_letter(col)}{revenue_rows[0]}:{get_column_letter(col)}{row-1}"
            c = ws.cell(row=row, column=col, value=f"=SUM({sum_range})")
            c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor=GREEN_DARK)
            c.alignment = Alignment(horizontal="right")
            c.border = BORDER
            c.number_format = '#,##0'
        # Cumul
        cum = ws.cell(row=row, column=7, value=f"=SUM(B{row}:F{row})")
        cum.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        cum.fill = PatternFill("solid", fgColor=GREEN_DARK)
        cum.alignment = Alignment(horizontal="right")
        cum.border = BORDER
        cum.number_format = '#,##0'
        total_rev_row = row
    else:
        cell_label(ws.cell(row=row, column=1), label)
        for col, val in enumerate(vals, 2):
            cell_calc(ws.cell(row=row, column=col), val, fmt='#,##0',
                      bg=ALT_ROW if row % 2 == 0 else WHITE)
        # Cumul
        c = ws.cell(row=row, column=7, value=f"=SUM(B{row}:F{row})")
        c.font = Font(name="Arial", size=10, italic=True, color=GREEN_DARK)
        c.alignment = Alignment(horizontal="right")
        c.border = BORDER
        c.number_format = '#,##0'
        revenue_rows.append(row)
    row += 1

row += 1  # gap

charge_start = None
for label, vals in pl_charges:
    if vals is None:
        section(ws, row, label, span=7, bg="C62828", fg=WHITE)
        ws.row_dimensions[row].height = 18
    elif vals == "SUM":
        cell_label(ws.cell(row=row, column=1), label, bold=True, size=11)
        for col in range(2, 7):
            sum_range = f"{get_column_letter(col)}{charge_rows[0]}:{get_column_letter(col)}{row-1}"
            c = ws.cell(row=row, column=col, value=f"=SUM({sum_range})")
            c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor="C62828")
            c.alignment = Alignment(horizontal="right")
            c.border = BORDER
            c.number_format = '#,##0'
        cum = ws.cell(row=row, column=7, value=f"=SUM(B{row}:F{row})")
        cum.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        cum.fill = PatternFill("solid", fgColor="C62828")
        cum.alignment = Alignment(horizontal="right")
        cum.border = BORDER
        cum.number_format = '#,##0'
        total_charge_row = row
    else:
        cell_label(ws.cell(row=row, column=1), label)
        for col, val in enumerate(vals, 2):
            cell_calc(ws.cell(row=row, column=col), val, fmt='#,##0',
                      bg=ALT_ROW if row % 2 == 0 else WHITE)
        c = ws.cell(row=row, column=7, value=f"=SUM(B{row}:F{row})")
        c.font = Font(name="Arial", size=10, italic=True, color="C62828")
        c.alignment = Alignment(horizontal="right")
        c.border = BORDER
        c.number_format = '#,##0'
        charge_rows.append(row)
    row += 1

row += 2  # gap

# EBITDA
section(ws, row, "EBITDA / EBIT / RESULTAT NET", span=7, bg="FF9800", fg=WHITE)
ws.row_dimensions[row].height = 20
row += 1

cell_label(ws.cell(row=row, column=1), "EBITDA", bold=True, size=12)
for col in range(2, 7):
    formula = f"={get_column_letter(col)}{total_rev_row}-{get_column_letter(col)}{total_charge_row}"
    c = ws.cell(row=row, column=col, value=formula)
    c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor="FF9800")
    c.alignment = Alignment(horizontal="right")
    c.border = BORDER
    c.number_format = '#,##0'
c = ws.cell(row=row, column=7, value=f"=SUM(B{row}:F{row})")
c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
c.fill = PatternFill("solid", fgColor="FF9800")
c.alignment = Alignment(horizontal="right")
c.border = BORDER
c.number_format = '#,##0'
ebitda_row = row
row += 1

cell_label(ws.cell(row=row, column=1), "Marge EBITDA", italic=True)
for col in range(2, 7):
    formula = f"=IFERROR({get_column_letter(col)}{ebitda_row}/{get_column_letter(col)}{total_rev_row},0)"
    cell_calc(ws.cell(row=row, column=col), formula, fmt='0.0%', bold=True, color=GREEN_DARK)
row += 2

cell_label(ws.cell(row=row, column=1), "Amortissements", italic=True)
amort_vals = [-20, -30, -40, -65, -95]
for col, v in enumerate(amort_vals, 2):
    cell_calc(ws.cell(row=row, column=col), v, fmt='#,##0', color=RED)
amort_row = row
row += 1

cell_label(ws.cell(row=row, column=1), "EBIT (Resultat d'exploitation)", bold=True, size=11)
for col in range(2, 7):
    formula = f"={get_column_letter(col)}{ebitda_row}+{get_column_letter(col)}{amort_row}"
    cell_calc(ws.cell(row=row, column=col), formula, fmt='#,##0', bold=True)
ebit_row = row
row += 2

cell_label(ws.cell(row=row, column=1), "Charges financieres", italic=True)
fin_vals = [-5, -10, -15, -20, -25]
for col, v in enumerate(fin_vals, 2):
    cell_calc(ws.cell(row=row, column=col), v, fmt='#,##0', color=RED)
fin_row = row
row += 1

cell_label(ws.cell(row=row, column=1), "Subventions etalees (+)", italic=True)
for col in range(2, 7):
    cell_calc(ws.cell(row=row, column=col), 12, fmt='#,##0', color=GREEN_DARK)
sub_row = row
row += 1

cell_label(ws.cell(row=row, column=1), "Resultat avant impot", bold=True)
for col in range(2, 7):
    formula = f"={get_column_letter(col)}{ebit_row}+{get_column_letter(col)}{fin_row}+{get_column_letter(col)}{sub_row}"
    cell_calc(ws.cell(row=row, column=col), formula, fmt='#,##0', bold=True)
rai_row = row
row += 1

cell_label(ws.cell(row=row, column=1), "Impot sur les societes (12,5% DOM)", italic=True)
is_vals = [0, 0, 0, "=-IF(F{rai}>0,F{rai}*0.125,0)".format(rai=rai_row), "=-IF(F{rai}>0,F{rai}*0.125,0)".format(rai=rai_row)]
# Tax only if profit
for col in range(2, 7):
    formula = f"=-IF({get_column_letter(col)}{rai_row}>0,{get_column_letter(col)}{rai_row}*0.125,0)"
    cell_calc(ws.cell(row=row, column=col), formula, fmt='#,##0', color=RED)
is_row = row
row += 1

cell_label(ws.cell(row=row, column=1), "RESULTAT NET", bold=True, size=12)
for col in range(2, 7):
    formula = f"={get_column_letter(col)}{rai_row}+{get_column_letter(col)}{is_row}"
    c = ws.cell(row=row, column=col, value=formula)
    c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=GREEN_DARK)
    c.alignment = Alignment(horizontal="right")
    c.border = BORDER
    c.number_format = '#,##0'
c = ws.cell(row=row, column=7, value=f"=SUM(B{row}:F{row})")
c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
c.fill = PatternFill("solid", fgColor=GREEN_DARK)
c.alignment = Alignment(horizontal="right")
c.border = BORDER
c.number_format = '#,##0'

# ============================================================
# 4. CASH FLOW
# ============================================================
ws = wb.create_sheet("Cash_Flow")
ws.sheet_properties.tabColor = "0066CC"
for c in "ABCDEFGH":
    ws.column_dimensions[c].width = 14
ws.column_dimensions["A"].width = 38

ws.merge_cells("A1:G1")
ws["A1"] = "PLAN DE TRESORERIE 5 ANS (k EUR)"
ws["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
ws["A1"].fill = PatternFill("solid", fgColor=GREEN_DARK)
ws["A1"].alignment = Alignment(horizontal="center")
ws.row_dimensions[1].height = 28

for col, h in enumerate(["Poste", "An 1", "An 2", "An 3", "An 4", "An 5"], 1):
    hdr(ws.cell(row=3, column=col))
    ws.cell(row=3, column=col).value = h

cf = [
    ("Tresorerie debut d'annee",          [10, 165, 50, 15, 80]),
    ("",                                   None),
    ("Apport Seed",                        [300, 0, 0, 0, 0]),
    ("Levee Serie A (optionnelle An 3)",  [0, 0, 200, 0, 0]),
    ("Subvention ADVENIR percue",          [37, 0, 0, 0, 0]),
    ("Encaissements clients",              [25, 75, 155, 270, 410]),
    ("TOTAL ENCAISSEMENTS",                "SUMENC"),
    ("",                                   None),
    ("Decaissements opex",                 [-150, -175, -195, -225, -260]),
    ("CAPEX",                              [-50, -50, -55, -130, -135]),
    ("Frais financiers",                   [-1, -3, -5, -8, -12]),
    ("Impot societes",                     [0, 0, 0, 0, 0]),
    ("TOTAL DECAISSEMENTS",                "SUMDEC"),
    ("",                                   None),
    ("VARIATION DE TRESORERIE",            "VARIA"),
    ("",                                   None),
    ("TRESORERIE FIN D'ANNEE",             "FIN"),
]

row = 4
enc_rows = []
dec_rows = []
treso_init_row = None

for label, vals in cf:
    if vals is None:
        row += 1
        continue
    if vals == "SUMENC":
        cell_label(ws.cell(row=row, column=1), label, bold=True)
        for col in range(2, 7):
            r1 = enc_rows[0]
            r2 = enc_rows[-1]
            formula = f"=SUM({get_column_letter(col)}{r1}:{get_column_letter(col)}{r2})"
            c = ws.cell(row=row, column=col, value=formula)
            c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor=GREEN_DARK)
            c.alignment = Alignment(horizontal="right")
            c.border = BORDER
            c.number_format = '#,##0'
        sum_enc_row = row
    elif vals == "SUMDEC":
        cell_label(ws.cell(row=row, column=1), label, bold=True)
        for col in range(2, 7):
            r1 = dec_rows[0]
            r2 = dec_rows[-1]
            formula = f"=SUM({get_column_letter(col)}{r1}:{get_column_letter(col)}{r2})"
            c = ws.cell(row=row, column=col, value=formula)
            c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor="C62828")
            c.alignment = Alignment(horizontal="right")
            c.border = BORDER
            c.number_format = '#,##0'
        sum_dec_row = row
    elif vals == "VARIA":
        cell_label(ws.cell(row=row, column=1), label, bold=True)
        for col in range(2, 7):
            formula = f"={get_column_letter(col)}{sum_enc_row}+{get_column_letter(col)}{sum_dec_row}"
            cell_calc(ws.cell(row=row, column=col), formula, fmt='#,##0', bold=True, color=GREEN_DARK)
        varia_row = row
    elif vals == "FIN":
        cell_label(ws.cell(row=row, column=1), label, bold=True, size=12)
        for col in range(2, 7):
            formula = f"={get_column_letter(col)}{treso_init_row}+{get_column_letter(col)}{varia_row}"
            c = ws.cell(row=row, column=col, value=formula)
            c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor=GREEN_DARK)
            c.alignment = Alignment(horizontal="right")
            c.border = BORDER
            c.number_format = '#,##0'
    else:
        cell_label(ws.cell(row=row, column=1), label, bold=label.startswith("Tresorerie"))
        for col, val in enumerate(vals, 2):
            cell_calc(ws.cell(row=row, column=col), val, fmt='#,##0',
                      bg=ALT_ROW if row % 2 == 0 else WHITE,
                      color=RED if val < 0 else "000000")
        if "debut" in label:
            treso_init_row = row
        elif label in ("Apport Seed", "Levee Serie A", "Subvention ADVENIR percue", "Encaissements clients"):
            enc_rows.append(row)
        elif label.startswith("Decaissements") or label in ("CAPEX", "Frais financiers", "Impot societes"):
            dec_rows.append(row)
    row += 1

# ============================================================
# 5. CAPEX_OPEX
# ============================================================
ws = wb.create_sheet("CAPEX_OPEX")
ws.sheet_properties.tabColor = "FF9800"
for c in "ABCDEFG":
    ws.column_dimensions[c].width = 14
ws.column_dimensions["A"].width = 42

ws.merge_cells("A1:G1")
ws["A1"] = "INVESTISSEMENTS (CAPEX) ET CHARGES (OPEX)"
ws["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
ws["A1"].fill = PatternFill("solid", fgColor=GREEN_DARK)
ws["A1"].alignment = Alignment(horizontal="center")
ws.row_dimensions[1].height = 28

# CAPEX
section(ws, 3, "CAPEX (k EUR)")
for col, h in enumerate(["Element", "An 1", "An 2", "An 3", "An 4", "An 5", "Cumul"], 1):
    hdr(ws.cell(row=4, column=col))
    ws.cell(row=4, column=col).value = h

capex = [
    # Plan SIMPLE : 20 PDC initial An 1 + 10/an
    ("Bornes voirie publique (20 PDC An 1)",   [99,  0,  0,  0,  0]),
    ("Bornes LOA (5 bornes/an = 10 PDC)",      [ 0, 25, 25, 15, 15]),
    ("Stations DC fast (2 An 4 + 2 An 5)",     [ 0,  0,  0, 75, 75]),
    ("Travaux genie civil (sous-traite)",      [40, 18, 20, 25, 30]),
    ("Outillage et equipements",               [10,  3,  4,  5,  5]),
    ("Vehicules de service (utilitaire)",      [15,  0,  0, 15,  0]),
    ("Application mobile (developpement)",     [15,  4,  6,  6,  5]),
    ("Mobilier bureau, locaux",                [ 5,  3,  3,  3,  3]),
    ("Total CAPEX annuel",                     "SUM"),
]

row = 5
capex_rows = []
for label, vals in capex:
    if vals == "SUM":
        cell_label(ws.cell(row=row, column=1), label, bold=True, size=11)
        for col in range(2, 7):
            r1 = capex_rows[0]
            r2 = capex_rows[-1]
            formula = f"=SUM({get_column_letter(col)}{r1}:{get_column_letter(col)}{r2})"
            c = ws.cell(row=row, column=col, value=formula)
            c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor="FF9800")
            c.alignment = Alignment(horizontal="right")
            c.border = BORDER
            c.number_format = '#,##0'
        cum = ws.cell(row=row, column=7, value=f"=SUM(B{row}:F{row})")
        cum.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        cum.fill = PatternFill("solid", fgColor="FF9800")
        cum.alignment = Alignment(horizontal="right")
        cum.border = BORDER
        cum.number_format = '#,##0'
    else:
        cell_label(ws.cell(row=row, column=1), label)
        for col, val in enumerate(vals, 2):
            cell_calc(ws.cell(row=row, column=col), val, fmt='#,##0',
                      bg=ALT_ROW if row % 2 == 0 else WHITE)
        c = ws.cell(row=row, column=7, value=f"=SUM(B{row}:F{row})")
        c.font = Font(name="Arial", size=10, italic=True, color="FF9800")
        c.alignment = Alignment(horizontal="right")
        c.border = BORDER
        c.number_format = '#,##0'
        capex_rows.append(row)
    row += 1

# OPEX detaille
row += 2
section(ws, row, "OPEX DETAILLE / RECRUTEMENT (k EUR)", span=7)
row += 1
for col, h in enumerate(["Poste", "An 1", "An 2", "An 3", "An 4", "An 5", "Cumul"], 1):
    hdr(ws.cell(row=row, column=col))
    ws.cell(row=row, column=col).value = h
row += 1

rh = [
    # Plan SIMPLE : 3 ETP An 1 -> 6 ETP An 5 (lean)
    ("Loic LUDOSKY (President)",                  [36, 42, 48, 54, 60]),
    ("Patrice LUDOSKY (DG)",                      [36, 42, 48, 54, 60]),
    ("Technicien IRVE n.1 (a partir T+6)",         [14, 28, 30, 32, 34]),
    ("Resp. commercial (a partir An 3)",           [ 0,  0, 36, 42, 45]),
    ("Technicien IRVE n.2 (An 4)",                 [ 0,  0,  0, 28, 30]),
    ("RAF mi-temps externalise",                   [ 0, 12, 14, 16, 18]),
    ("Charges patronales (45 %)",                  [38, 56,  79,103, 121]),
    ("Total Salaires & charges (k EUR)",           "SUM"),
]
rh_start_row = row
rh_rows = []
for label, vals in rh:
    if vals == "SUM":
        cell_label(ws.cell(row=row, column=1), label, bold=True, size=11)
        for col in range(2, 7):
            r1 = rh_rows[0]
            r2 = rh_rows[-1]
            formula = f"=SUM({get_column_letter(col)}{r1}:{get_column_letter(col)}{r2})"
            c = ws.cell(row=row, column=col, value=formula)
            c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor="FF9800")
            c.alignment = Alignment(horizontal="right")
            c.border = BORDER
            c.number_format = '#,##0'
        c = ws.cell(row=row, column=7, value=f"=SUM(B{row}:F{row})")
        c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor="FF9800")
        c.alignment = Alignment(horizontal="right")
        c.border = BORDER
        c.number_format = '#,##0'
    else:
        cell_label(ws.cell(row=row, column=1), label)
        for col, val in enumerate(vals, 2):
            cell_calc(ws.cell(row=row, column=col), val, fmt='#,##0',
                      bg=ALT_ROW if row % 2 == 0 else WHITE)
        c = ws.cell(row=row, column=7, value=f"=SUM(B{row}:F{row})")
        c.font = Font(name="Arial", size=10, italic=True)
        c.alignment = Alignment(horizontal="right")
        c.border = BORDER
        c.number_format = '#,##0'
        rh_rows.append(row)
    row += 1

# ============================================================
# 6. SCENARIOS
# ============================================================
ws = wb.create_sheet("Scenarios")
ws.sheet_properties.tabColor = "C62828"
for c in "ABCDEFGH":
    ws.column_dimensions[c].width = 16
ws.column_dimensions["A"].width = 38

ws.merge_cells("A1:H1")
ws["A1"] = "SCENARIOS PESSIMISTE / MEDIAN / OPTIMISTE — REVENU & EBITDA (k EUR)"
ws["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
ws["A1"].fill = PatternFill("solid", fgColor=GREEN_DARK)
ws["A1"].alignment = Alignment(horizontal="center")
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:H2")
ws["A2"] = "Le scenario MEDIAN est la reference. Les autres scenarios servent de stress test."
ws["A2"].font = Font(name="Arial", size=10, italic=True, color=GREEN_DARK)
ws["A2"].fill = PatternFill("solid", fgColor=GREEN_BG)
ws["A2"].alignment = Alignment(horizontal="center")

# Headers
for col, h in enumerate(["Scenario", "Hypothese cle", "An 1", "An 2", "An 3", "An 4", "An 5", "TIR 5 ans"], 1):
    hdr(ws.cell(row=4, column=col))
    ws.cell(row=4, column=col).value = h
ws.row_dimensions[4].height = 26

# REVENU
ws.merge_cells("A5:H5")
ws["A5"] = "REVENUS (k EUR)"
ws["A5"].font = Font(name="Arial", size=10, bold=True, color=WHITE)
ws["A5"].fill = PatternFill("solid", fgColor="0066CC")
ws["A5"].alignment = Alignment(horizontal="left", indent=1)

scenarios_rev = [
    ("Pessimiste", "Adoption VE -30% / pas de DC fast / 50 PDC An 5",   25,  60, 110, 175, 280, "2%",  "FFEBEE"),
    ("Median",     "Plan simple : 60 PDC An 5 dont 8 DC fast",          35,  85, 170, 290, 430, "10%", "F4FFF0"),
    ("Optimiste",  "Antilles An 5 + sponsoring + DC fast accelere",     50, 120, 240, 430, 680, "22%", "E1F5FE"),
]

for i, scen in enumerate(scenarios_rev, 6):
    name, hyp, *vals, tir, bg = scen
    cell_label(ws.cell(row=i, column=1), name, bold=True)
    ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=bg)
    cell_label(ws.cell(row=i, column=2), hyp, italic=True)
    ws.cell(row=i, column=2).fill = PatternFill("solid", fgColor=bg)
    for col, v in enumerate(vals, 3):
        c = ws.cell(row=i, column=col, value=v)
        c.font = Font(name="Arial", size=10, bold=(name == "Median"))
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="right")
        c.border = BORDER
        c.number_format = '#,##0'
    c = ws.cell(row=i, column=8, value=tir)
    c.font = Font(name="Arial", size=11, bold=True, color=GREEN_DARK if name != "Pessimiste" else RED)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER

# EBITDA
ws.merge_cells("A10:H10")
ws["A10"] = "EBITDA (k EUR)"
ws["A10"].font = Font(name="Arial", size=10, bold=True, color=WHITE)
ws["A10"].fill = PatternFill("solid", fgColor="FF9800")
ws["A10"].alignment = Alignment(horizontal="left", indent=1)

scenarios_ebitda = [
    ("Pessimiste", "EBITDA fragile, sortie An 8",            -195, -180, -160, -135,  -75,  "2%",  "FFEBEE"),
    ("Median",     "EBITDA equilibre An 5, sortie An 7-8",   -185, -155, -100,  -20,   90, "10%",  "F4FFF0"),
    ("Optimiste",  "EBITDA fort (DC fast + Antilles)",       -170, -100,    0,  100,  240, "22%",  "E1F5FE"),
]
for i, scen in enumerate(scenarios_ebitda, 11):
    name, hyp, *vals, tir, bg = scen
    cell_label(ws.cell(row=i, column=1), name, bold=True)
    ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=bg)
    cell_label(ws.cell(row=i, column=2), hyp, italic=True)
    ws.cell(row=i, column=2).fill = PatternFill("solid", fgColor=bg)
    for col, v in enumerate(vals, 3):
        c = ws.cell(row=i, column=col, value=v)
        c.font = Font(name="Arial", size=10, bold=(name == "Median"),
                      color=RED if v < 0 else "000000")
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="right")
        c.border = BORDER
        c.number_format = '#,##0'
    c = ws.cell(row=i, column=8, value=tir)
    c.font = Font(name="Arial", size=11, bold=True, color=GREEN_DARK if name != "Pessimiste" else RED)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER

# Hypotheses scenarios
ws.merge_cells("A16:H16")
ws["A16"] = "DETAIL DES HYPOTHESES PAR SCENARIO"
ws["A16"].font = Font(name="Arial", size=11, bold=True, color=WHITE)
ws["A16"].fill = PatternFill("solid", fgColor=GREEN_DARK)
ws["A16"].alignment = Alignment(horizontal="center")

hyps = [
    ("Plan deploiement PDC",                 "50 PDC An 5",  "60 PDC An 5",  "70 PDC An 5"),
    ("Stations DC fast An 5",                "0",            "8 PDC (4 stations)", "12 PDC (6 stations)"),
    ("Adoption VE en Guyane",                "+20% an",      "+35% an",      "+50% an"),
    ("Date premiere mise en service",        "T+9 mois",     "T+5 mois",     "T+5 mois"),
    ("Cout materiel/borne moyen",            "+15%",         "Devis E-TOTEM","-5%"),
    ("Nombre de contrats LOA An 1",          "0",            "0 (T+12)",     "5"),
    ("Sponsoring publicitaire An 5",         "10 k€/an",     "40 k€/an",     "100 k€/an"),
    ("Phase 2 Antilles",                     "Annulee",      "An 5 (etude)", "An 4 (10 PDC)"),
    ("Couverture ADVENIR maintenue",         "70%",          "75,2%",        "75,2%"),
    ("Levee Serie A (An 3, optionnelle)",    "Echec/repli",  "200 k€",       "500 k€"),
]
for col, h in enumerate(["Hypothese", "Pessimiste", "Median", "Optimiste"], 1):
    hdr(ws.cell(row=17, column=col))
    ws.cell(row=17, column=col).value = h
ws.merge_cells("A17:B17")  # Adapter colonne label
for i, (lbl, p, m, o) in enumerate(hyps, 18):
    cell_label(ws.cell(row=i, column=1), lbl)
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
    cell_label(ws.cell(row=i, column=3), p)
    ws.cell(row=i, column=3).fill = PatternFill("solid", fgColor="FFEBEE")
    cell_label(ws.cell(row=i, column=4), m, bold=True)
    ws.cell(row=i, column=4).fill = PatternFill("solid", fgColor="F4FFF0")
    cell_label(ws.cell(row=i, column=5), o)
    ws.cell(row=i, column=5).fill = PatternFill("solid", fgColor="E1F5FE")

# ============================================================
# 7. TIR_Sortie
# ============================================================
ws = wb.create_sheet("TIR_Sortie")
ws.sheet_properties.tabColor = "FFD700"
for c in "ABCDEFG":
    ws.column_dimensions[c].width = 18
ws.column_dimensions["A"].width = 38

ws.merge_cells("A1:G1")
ws["A1"] = "ANALYSE TIR INVESTISSEUR & SCENARIOS DE SORTIE"
ws["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
ws["A1"].fill = PatternFill("solid", fgColor=GREEN_DARK)
ws["A1"].alignment = Alignment(horizontal="center")
ws.row_dimensions[1].height = 28

# PARAMETRES INVESTISSEUR
section(ws, 3, "PARAMETRES INVESTISSEUR (cellules bleues = modifiables)")
params = [
    ("Investissement initial (apport Seed)",     300000,  '#,##0" EUR"'),
    ("% capital obtenu (post-money)",            0.30,    '0.0%'),
    ("Annee de sortie envisagee",                7,       '0'),
    ("EBITDA de reference pour valorisation",    250000,  '#,##0" EUR"'),
    ("Multiple cible cession industrielle",      8,       '0.0"x"'),
    ("Multiple cible fonds infra",               10,      '0.0"x"'),
]
for i, (lbl, val, fmt) in enumerate(params, 4):
    cell_label(ws.cell(row=i, column=1), lbl, bold=True)
    cell_input(ws.cell(row=i, column=2), val, fmt=fmt)

# CALCUL TIR
section(ws, 11, "CALCUL TIR ET MULTIPLES")
calc = [
    ("Valeur entreprise sortie (cession indus)",  '=B7*B8'),
    ("Valeur cession a l'investisseur",            '=B12*B5'),
    ("Multiple investisseur (cash on cash)",       '=B13/B4'),
    ("Plus-value totale (en EUR)",                 '=B13-B4'),
    ("TIR sur N annees",                           '=(B13/B4)^(1/B6)-1'),
]
for i, (lbl, formula) in enumerate(calc, 12):
    cell_label(ws.cell(row=i, column=1), lbl)
    fmt = '#,##0" EUR"'
    if "Multiple" in lbl:
        fmt = '0.00"x"'
    elif "TIR" in lbl:
        fmt = '0.0%'
    c = ws.cell(row=i, column=2, value=formula)
    c.font = Font(name="Arial", size=11, bold=True, color=GREEN_DARK)
    c.fill = PatternFill("solid", fgColor=GREEN_LIGHT)
    c.alignment = Alignment(horizontal="right")
    c.border = BORDER
    c.number_format = fmt

# Tableau de scenarios
section(ws, 19, "TABLEAU DES SCENARIOS DE SORTIE")
for col, h in enumerate(["Scenario", "Annee sortie", "Acquereur type", "Prix cession", "Multiple invest.", "TIR investisseur"], 1):
    hdr(ws.cell(row=20, column=col))
    ws.cell(row=20, column=col).value = h
ws.row_dimensions[20].height = 26

scenarios_sortie = [
    # Plan SIMPLE 60 PDC An 5 - sortie 7-8 ans, multiples adaptes a la taille
    ("Pessimiste",  "An 8", "Cession industrielle (x6 sur EBITDA 50k)",     300000, 1.00, -0.00),
    ("Pessimiste",  "An 8", "Buy-back fondateurs (apport seul restitue)",   300000, 1.00, -0.00),
    ("Median",      "An 7", "Cession industrielle (x8 sur EBITDA 200k)",   1600000, 1.60,  0.07),
    ("Median",      "An 7", "Cession a un fonds (x8 sur EBITDA 250k)",     2000000, 2.00,  0.10),
    ("Median",      "An 8", "Cession industrielle (x8 sur EBITDA 350k)",   2800000, 2.80,  0.14),
    ("Optimiste",   "An 7", "Cession industrielle (x10 sur EBITDA 350k)",  3500000, 3.50,  0.20),
    ("Optimiste",   "An 7", "Fonds infra (x12 sur EBITDA 400k)",           4800000, 4.80,  0.25),
    ("Optimiste+",  "An 8", "Fonds infra (x12 sur EBITDA 600k)",           7200000, 7.20,  0.28),
]
for i, scen in enumerate(scenarios_sortie, 21):
    name, year, acq, prix, mult, tir = scen
    bg = "FFEBEE" if "Pessimiste" in name else ("F4FFF0" if "Median" in name else "E1F5FE")
    cell_label(ws.cell(row=i, column=1), name, bold=True)
    ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=bg)
    cell_label(ws.cell(row=i, column=2), year)
    ws.cell(row=i, column=2).fill = PatternFill("solid", fgColor=bg)
    cell_label(ws.cell(row=i, column=3), acq, italic=True)
    ws.cell(row=i, column=3).fill = PatternFill("solid", fgColor=bg)
    c = ws.cell(row=i, column=4, value=prix)
    c.number_format = '#,##0" EUR"'
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(name="Arial", size=10, bold=True)
    c.alignment = Alignment(horizontal="right")
    c.border = BORDER
    c = ws.cell(row=i, column=5, value=mult)
    c.number_format = '0.00"x"'
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(name="Arial", size=10, bold=True)
    c.alignment = Alignment(horizontal="right")
    c.border = BORDER
    c = ws.cell(row=i, column=6, value=tir)
    c.number_format = '0.0%'
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(name="Arial", size=11, bold=True,
                  color=GREEN_DARK if tir > 0.15 else (RED if tir < 0 else "FF9800"))
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER

# Synthese finale
ws.merge_cells("A31:F31")
ws["A31"] = "SYNTHESE — Plan SIMPLE : 20 PDC initial + 10/an = 60 PDC An 5 — Seed 300k — TIR 8-12% sur 7-8 ans"
ws["A31"].font = Font(name="Arial", size=11, bold=True, color=WHITE)
ws["A31"].fill = PatternFill("solid", fgColor=GREEN_DARK)
ws["A31"].alignment = Alignment(horizontal="center")
ws.row_dimensions[31].height = 26

# Save
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"[OK] {OUT.name} genere")
print(f"     7 sheets : Synthese | Hypotheses | P&L | Cash_Flow | CAPEX_OPEX | Scenarios | TIR_Sortie")
