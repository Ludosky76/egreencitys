"""
Genere Cibles_Entreprises_Guyane.xlsx
  Sheet 1 : "Cibles"        — ~50 entreprises cibles LOA en Guyane
  Sheet 2 : "Calculateur"   — simulateur LOA interactif
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "_dossiers" / "03_Offre_Entreprises_LOA" / "Cibles_Entreprises_Guyane.xlsx"

wb = Workbook()

# ============================================================
# SHEET 1 — CIBLES
# ============================================================
ws = wb.active
ws.title = "Cibles"

# Styles
GREEN_DARK  = "1a472a"
GREEN_MED   = "2e7d32"
GREEN_LIGHT = "e8f5e9"
HEADER_BG   = "1a472a"
HEADER_FG   = "FFFFFF"
ALT_ROW     = "f1f8e9"
ORANGE      = "f57c00"
GREY        = "607d8b"

def hdr_style(cell, bg=HEADER_BG, fg=HEADER_FG, size=10, bold=True):
    cell.font = Font(name="Arial", size=size, bold=bold, color=fg)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def border_thin():
    s = Side(style="thin", color="c8e6c9")
    return Border(left=s, right=s, top=s, bottom=s)

# Title row
ws.merge_cells("A1:R1")
ws["A1"] = "EGREENCITY'S — CIBLES COMMERCIALES LOA BORNES DE RECHARGE — GUYANE 2026"
ws["A1"].font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor=GREEN_DARK)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Sub-title
ws.merge_cells("A2:R2")
ws["A2"] = "Programme ADVENIR EGREENCITY'S | LOA 72 mois min. | A partir de 89 EUR HT/mois/PDC | egreencitys@gmail.com | +33 6 51 14 11 18"
ws["A2"].font = Font(name="Arial", size=9, italic=True, color=GREEN_DARK)
ws["A2"].fill = PatternFill("solid", fgColor="c8e6c9")
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

# Column headers
HEADERS = [
    ("ID",           4),
    ("Secteur",     16),
    ("Entreprise",  26),
    ("Ville",       14),
    ("Adresse",     30),
    ("Tel",         16),
    ("Email",       30),
    ("Contact",     20),
    ("Offre LOA",   14),
    ("PDC",          5),
    ("Loyer/mois HT",14),
    ("Statut",      14),
    ("Date contact", 12),
    ("Reponse",     12),
    ("RDV",         12),
    ("Contrat",     12),
    ("Notes",       35),
    ("Priorite",    10),
]

ws.row_dimensions[3].height = 32
for col, (hdr, width) in enumerate(HEADERS, 1):
    c = ws.cell(row=3, column=col, value=hdr)
    hdr_style(c)
    c.border = border_thin()
    ws.column_dimensions[get_column_letter(col)].width = width

# Données entreprises cibles
TARGETS = [
    # (ID, Secteur, Entreprise, Ville, Adresse, Tel, Email, Contact, Offre, PDC, Loyer, Priorite)
    # HOTELS
    ("H01", "Hotellerie", "Hotel Mercure Cayenne", "Cayenne",
     "Place des Palmistes, 97300 Cayenne",
     "+33 5 94 29 20 20", "h3473@accor.com", "Direction",
     "Premium", 2, 338, "Priorite 1"),
    ("H02", "Hotellerie", "Best Western Hotel le Dronmi", "Cayenne",
     "Chemin St-Antoine, 97300 Cayenne",
     "+33 5 94 25 06 06", "", "Direction",
     "Pro", 2, 278, "Priorite 1"),
    ("H03", "Hotellerie", "Hotel Novotel Cayenne", "Cayenne",
     "Route de Raban, 97300 Cayenne",
     "+33 5 94 30 19 00", "", "Direction",
     "Premium", 2, 338, "Priorite 1"),
    ("H04", "Hotellerie", "Hotel Atlantis Kourou", "Kourou",
     "Boulevard du Bagne, 97310 Kourou",
     "+33 5 94 32 13 00", "", "Direction",
     "Pro", 2, 278, "Priorite 2"),
    ("H05", "Hotellerie", "Auberge des Orchidees", "Saint-Laurent-du-Maroni",
     "Route Nationale 1, 97320 SLM",
     "+33 5 94 34 00 21", "", "Direction",
     "Standard", 2, 238, "Priorite 3"),
    ("H06", "Hotellerie", "Hotel Le Grenat Mana", "Mana",
     "Centre bourg, 97318 Mana",
     "", "", "Direction",
     "Starter", 1, 89, "Priorite 3"),
    # GRANDE DISTRIBUTION
    ("D01", "Grande distribution", "Geant Casino Cayenne", "Cayenne",
     "Avenue Louis-Constant Fleming, 97300 Cayenne",
     "+33 5 94 25 72 00", "", "Responsable commercial",
     "Premium", 4, 676, "Priorite 1"),
    ("D02", "Grande distribution", "Carrefour Galeria Cayenne", "Cayenne",
     "Galeria Shopping Center, 97300 Cayenne",
     "+33 5 94 29 65 00", "", "Responsable commercial",
     "Premium", 4, 676, "Priorite 1"),
    ("D03", "Grande distribution", "Family Plaza Matoury", "Matoury",
     "Zone commerciale, 97351 Matoury",
     "+33 5 94 35 86 00", "", "Direction",
     "Pro", 2, 278, "Priorite 1"),
    ("D04", "Grande distribution", "Hyper U Remire-Montjoly", "Remire-Montjoly",
     "Route de Montjoly, 97354 Remire-Montjoly",
     "", "", "Direction",
     "Pro", 2, 278, "Priorite 2"),
    ("D05", "Grande distribution", "Leader Price Kourou", "Kourou",
     "Cite Mirza, 97310 Kourou",
     "", "", "Direction",
     "Standard", 2, 238, "Priorite 2"),
    # AUTOMOBILE
    ("A01", "Automobile", "Sodirep (Renault-Dacia)", "Cayenne",
     "Route de Baduel, 97300 Cayenne",
     "+33 5 94 31 09 09", "", "Chef des ventes",
     "Premium", 4, 676, "Priorite 1"),
    ("A02", "Automobile", "Polygone (Peugeot-Citroen)", "Cayenne",
     "Route Nationale 1, 97300 Cayenne",
     "+33 5 94 25 58 00", "", "Chef des ventes",
     "Premium", 4, 676, "Priorite 1"),
    ("A03", "Automobile", "Toyota Guyane", "Cayenne",
     "Route de la Chaumiere, 97300 Cayenne",
     "+33 5 94 38 00 00", "", "Direction",
     "Pro", 2, 278, "Priorite 1"),
    ("A04", "Automobile", "Volkswagen Guyane (GDS)", "Cayenne",
     "Route de Baduel, 97300 Cayenne",
     "", "", "Direction",
     "Pro", 2, 278, "Priorite 2"),
    ("A05", "Automobile", "Hyundai Guyane", "Remire-Montjoly",
     "Route de l'Arsenal, 97354 Remire-Montjoly",
     "", "", "Direction",
     "Standard", 2, 238, "Priorite 2"),
    ("A06", "Automobile", "BYD / Tesla reseller Guyane", "Cayenne",
     "(a identifier)", "", "", "Direction",
     "Pro", 2, 278, "Priorite 1"),
    # LOCATION VEHICULES
    ("L01", "Location vehicules", "Avis Car Rental Guyane", "Cayenne",
     "Aeroport Felix Eboue, 97351 Matoury",
     "+33 5 94 35 61 40", "", "Responsable agence",
     "Standard", 4, 476, "Priorite 1"),
    ("L02", "Location vehicules", "Europcar Guyane", "Cayenne",
     "Aeroport Felix Eboue, 97351 Matoury",
     "+33 5 94 35 63 13", "", "Responsable agence",
     "Standard", 4, 476, "Priorite 1"),
    ("L03", "Location vehicules", "Hertz Guyane", "Cayenne",
     "Aeroport Felix Eboue, 97351 Matoury",
     "+33 5 94 35 64 64", "", "Responsable agence",
     "Standard", 4, 476, "Priorite 1"),
    ("L04", "Location vehicules", "Sixt Guyane", "Cayenne",
     "Zone commerciale Degrad des Cannes",
     "", "", "Responsable agence",
     "Standard", 2, 238, "Priorite 2"),
    ("L05", "Location vehicules", "ADA Location Guyane", "Cayenne",
     "Route de Baduel, 97300 Cayenne",
     "", "", "Responsable",
     "Starter", 2, 178, "Priorite 2"),
    # BANQUES
    ("B01", "Banque / Assurance", "Credit Agricole Guyane", "Cayenne",
     "5 avenue du General de Gaulle, 97300 Cayenne",
     "+33 5 94 29 29 29", "contact@caguyane.fr", "Directeur regional",
     "Pro", 2, 278, "Priorite 2"),
    ("B02", "Banque / Assurance", "BRED Banque Populaire Guyane", "Cayenne",
     "2 place des Palmistes, 97300 Cayenne",
     "+33 5 94 30 10 00", "", "Direction agence",
     "Standard", 2, 238, "Priorite 2"),
    ("B03", "Banque / Assurance", "Societe Generale Guyane", "Cayenne",
     "Place Schoelcher, 97300 Cayenne",
     "+33 5 94 31 20 20", "", "Direction agence",
     "Standard", 2, 238, "Priorite 2"),
    ("B04", "Banque / Assurance", "BNC Guyane", "Cayenne",
     "Cite Mirza, 97300 Cayenne",
     "", "", "Direction",
     "Starter", 2, 178, "Priorite 3"),
    # SANTE
    ("S01", "Sante", "CHAR — Centre Hosp. Andrée Rosemon", "Cayenne",
     "Rue des Flamboyants, 97300 Cayenne",
     "+33 5 94 39 50 50", "", "Direction technique",
     "Pro", 4, 556, "Priorite 2"),
    ("S02", "Sante", "CHK — Centre Hosp. de Kourou", "Kourou",
     "Quartier Cogneau, 97310 Kourou",
     "+33 5 94 22 42 42", "", "Direction technique",
     "Standard", 2, 238, "Priorite 2"),
    ("S03", "Sante", "Hopital de Saint-Laurent-du-Maroni", "Saint-Laurent-du-Maroni",
     "Avenue du General de Gaulle, 97320 SLM",
     "+33 5 94 34 00 11", "", "Direction technique",
     "Standard", 2, 238, "Priorite 3"),
    ("S04", "Sante", "Clinique Veronique Cayenne", "Cayenne",
     "Route de Baduel, 97300 Cayenne",
     "", "", "Direction",
     "Standard", 2, 238, "Priorite 2"),
    # CSG / SPATIAL
    ("C01", "Spatial / Institutionnel", "CNES — Centre Spatial Guyanais", "Kourou",
     "Boulevard du Bagne, 97310 Kourou",
     "+33 5 94 33 51 11", "", "Responsable patrimoine",
     "Premium", 6, 1014, "Priorite 1"),
    ("C02", "Spatial / Institutionnel", "ArianeGroup Guyane", "Kourou",
     "CSG Kourou, 97310 Kourou",
     "", "", "Responsable",
     "Premium", 4, 676, "Priorite 1"),
    ("C03", "Spatial / Institutionnel", "ESA — European Space Agency Guyane", "Kourou",
     "CSG Kourou, 97310 Kourou",
     "", "", "Facilities Manager",
     "Pro", 4, 556, "Priorite 2"),
    ("C04", "Spatial / Institutionnel", "Avio (ancien ELV) Kourou", "Kourou",
     "CSG Kourou, 97310 Kourou",
     "", "", "Responsable",
     "Pro", 2, 278, "Priorite 2"),
    # STATIONS-SERVICE
    ("P01", "Energie / Carburant", "Sara Petroles Guyane — Siege", "Cayenne",
     "Zone industrielle, 97300 Cayenne",
     "+33 5 94 25 34 00", "", "Direction commerciale",
     "Premium", 4, 676, "Priorite 1"),
    ("P02", "Energie / Carburant", "TotalEnergies Cayenne (station flagship)", "Cayenne",
     "Route de Baduel, 97300 Cayenne",
     "", "", "Responsable station",
     "Premium", 2, 338, "Priorite 1"),
    ("P03", "Energie / Carburant", "Shell Guyane / EDF SEI", "Cayenne",
     "(a identifier)", "", "", "Direction",
     "Premium", 2, 338, "Priorite 2"),
    # TELECOM
    ("T01", "Telecom", "Orange Guyane", "Cayenne",
     "27 avenue du General de Gaulle, 97300 Cayenne",
     "+33 5 94 39 80 00", "", "Directeur regional",
     "Pro", 2, 278, "Priorite 2"),
    ("T02", "Telecom", "SFR Outremer Guyane", "Cayenne",
     "Centre commercial, 97300 Cayenne",
     "", "", "Direction agence",
     "Standard", 2, 238, "Priorite 2"),
    ("T03", "Telecom", "Digicel Guyane", "Cayenne",
     "Rue Lallouette, 97300 Cayenne",
     "", "", "Direction",
     "Standard", 2, 238, "Priorite 3"),
    # BTP
    ("E01", "BTP / Industrie", "Eiffage Guyane", "Remire-Montjoly",
     "Zone d'activite l'Arsenal, 97354 Remire-Montjoly",
     "+33 5 94 38 68 68", "", "Directeur technique",
     "Pro", 2, 278, "Priorite 2"),
    ("E02", "BTP / Industrie", "Bouygues Bâtiment DOM", "Cayenne",
     "Route de Montabo, 97300 Cayenne",
     "", "", "Direction locale",
     "Pro", 2, 278, "Priorite 2"),
    ("E03", "BTP / Industrie", "Vinci Construction DROM", "Cayenne",
     "Zone industrielle, 97300 Cayenne",
     "", "", "Direction locale",
     "Standard", 2, 238, "Priorite 2"),
    # AEROPORT / PORT
    ("I01", "Infrastructure", "Aeroport Felix Eboue (UAG)", "Matoury",
     "Route de l'Aeroport, 97351 Matoury",
     "+33 5 94 29 86 00", "", "Responsable concessions",
     "Premium", 6, 1014, "Priorite 1"),
    ("I02", "Infrastructure", "Port de Degrad-des-Cannes", "Remire-Montjoly",
     "Degrad-des-Cannes, 97354 Remire-Montjoly",
     "+33 5 94 29 95 10", "", "Direction exploitation",
     "Pro", 2, 278, "Priorite 2"),
    # GRANDE ENTREPRISE / MINES
    ("M01", "Industrie / Mines", "Montagne d'Or (Nord Resources)", "Cayenne",
     "Siege Cayenne — site a Paul-Isnard",
     "", "", "Direction RSE",
     "Pro", 2, 278, "Priorite 3"),
    ("M02", "Industrie / Mines", "CMB Metallurgie", "Cayenne",
     "Zone industrielle, 97300 Cayenne",
     "", "", "Direction",
     "Starter", 2, 178, "Priorite 3"),
    # ADMINISTRATION / COLLECTIVITE
    ("G01", "Administration", "Collectivite Territoriale de Guyane (CTG)", "Cayenne",
     "Cité administrative, 97300 Cayenne",
     "+33 5 94 29 20 20", "contact@ctguyane.fr", "DGA Transitions",
     "Premium", 4, 676, "Priorite 1"),
    ("G02", "Administration", "DEAL Guyane", "Cayenne",
     "258 rue de Limersack, 97300 Cayenne",
     "+33 5 94 29 75 00", "", "Direction",
     "Standard", 2, 238, "Priorite 3"),
]

# Statut vide par defaut
STATUS_OPTS = ""
DATE_EMPTY = ""
NOTES_EMPTY = "A contacter"

OFFRE_COLORS = {
    "Starter": "fff9c4",
    "Standard": "c8e6c9",
    "Pro": "bbdefb",
    "Premium": "ce93d8",
}
PRIO_COLORS = {
    "Priorite 1": "ffcdd2",
    "Priorite 2": "fff9c4",
    "Priorite 3": "f5f5f5",
}

for i, t in enumerate(TARGETS):
    row = i + 4
    ws.row_dimensions[row].height = 16
    fill_bg = ALT_ROW if i % 2 == 0 else "FFFFFF"
    id_, sect, ent, ville, adr, tel, email, contact, offre, pdc, loyer, prio = t
    values = [id_, sect, ent, ville, adr, tel, email, contact, offre, pdc, loyer,
              "", "", "", "", "", NOTES_EMPTY, prio]
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(vertical="center", wrap_text=(col == 5 or col == 17))
        c.border = border_thin()
        # Coloring
        if col == 9:  # Offre LOA
            bg = OFFRE_COLORS.get(str(val), fill_bg)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(name="Arial", size=9, bold=True)
        elif col == 18:  # Priorite
            bg = PRIO_COLORS.get(str(val), fill_bg)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(name="Arial", size=9, bold=True, color="333333")
        elif col == 11:  # Loyer
            c.number_format = '#,##0" EUR"'
            c.font = Font(name="Arial", size=9, bold=True, color=GREEN_MED)
        elif col == 6 and not val:  # Tel vide
            c.value = "(a completer)"
            c.font = Font(name="Arial", size=8, italic=True, color=GREY)
        elif col == 7 and not val:  # Email vide
            c.value = "(a completer)"
            c.font = Font(name="Arial", size=8, italic=True, color=GREY)
        else:
            c.fill = PatternFill("solid", fgColor=fill_bg)

# Freeze panes
ws.freeze_panes = "A4"

# Auto-filter
ws.auto_filter.ref = f"A3:R{3 + len(TARGETS)}"

# Legend row
legend_row = 4 + len(TARGETS) + 2
ws.merge_cells(f"A{legend_row}:D{legend_row}")
ws[f"A{legend_row}"] = "LEGENDE OFFRES LOA :"
ws[f"A{legend_row}"].font = Font(name="Arial", size=9, bold=True)
for col, (name, color) in enumerate([
    ("Starter (89 EUR)", "fff9c4"),
    ("Standard (119 EUR)", "c8e6c9"),
    ("Pro (139 EUR/PDC)", "bbdefb"),
    ("Premium (169 EUR/PDC)", "ce93d8"),
], 5):
    c = ws.cell(row=legend_row, column=col, value=name)
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(name="Arial", size=9, bold=True)
    c.alignment = Alignment(horizontal="center")

# ============================================================
# SHEET 2 — CALCULATEUR LOA
# ============================================================
wc = wb.create_sheet("Calculateur_LOA")
wc.sheet_view.showGridLines = True

wc.column_dimensions["A"].width = 28
wc.column_dimensions["B"].width = 22
wc.column_dimensions["C"].width = 22
wc.column_dimensions["D"].width = 22
wc.column_dimensions["E"].width = 22
wc.column_dimensions["F"].width = 22

# Title
wc.merge_cells("A1:F1")
wc["A1"] = "CALCULATEUR LOA BORNES — EGREENCITY'S 2026"
wc["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
wc["A1"].fill = PatternFill("solid", fgColor=GREEN_DARK)
wc["A1"].alignment = Alignment(horizontal="center", vertical="center")
wc.row_dimensions[1].height = 32

wc.merge_cells("A2:F2")
wc["A2"] = "Tous les montants sont en EUR HT — TVA Guyane 8,5 % applicable"
wc["A2"].font = Font(name="Arial", size=9, italic=True, color=GREEN_DARK)
wc["A2"].fill = PatternFill("solid", fgColor="c8e6c9")
wc["A2"].alignment = Alignment(horizontal="center")
wc.row_dimensions[2].height = 16

# Section 1 : Paramètres
wc.merge_cells("A4:F4")
wc["A4"] = "PARAMETRES DE SIMULATION (modifiez les cellules en bleu)"
wc["A4"].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
wc["A4"].fill = PatternFill("solid", fgColor="0000CC")
wc["A4"].alignment = Alignment(horizontal="left")
wc.row_dimensions[4].height = 20

params = [
    ("Offre selectionnee", "Pro", "B5",
     "Starter | Standard | Pro | Premium"),
    ("Nombre de PDC", 2, "B6",
     "1, 2, 4, 6, 8 ..."),
    ("Duree du contrat (mois)", 72, "B7",
     "72 a 96 mois"),
    ("Options supplementaires HT", 0, "B8",
     "Total des options (ecran, TPE, modem 4G...)"),
]

LOYERS_BASE = {
    "Starter": 89,
    "Standard": 119,
    "Pro": 139,
    "Premium": 169,
}
VAL_MAT = {
    "Starter": 715,
    "Standard": 1599,
    "Pro": 2739,
    "Premium": 4948,
}
OPT_ACHAT_PCT = 0.10
TVA_DOM = 0.085

for r, (label, default, cell_ref, note) in enumerate(params, 5):
    wc[f"A{r}"] = label
    wc[f"A{r}"].font = Font(name="Arial", size=10)
    wc[f"A{r}"].alignment = Alignment(vertical="center")
    c = wc[cell_ref]
    c.value = default
    c.font = Font(name="Arial", size=11, bold=True, color="0000FF")
    c.fill = PatternFill("solid", fgColor="dce6ff")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(
        left=Side(style="medium", color="0000CC"),
        right=Side(style="medium", color="0000CC"),
        top=Side(style="medium", color="0000CC"),
        bottom=Side(style="medium", color="0000CC"),
    )
    note_cell = wc[f"C{r}"]
    note_cell.value = f"→ {note}"
    note_cell.font = Font(name="Arial", size=8, italic=True, color=GREY)
    wc.merge_cells(f"C{r}:F{r}")

# Section 2 : Résultats
wc.merge_cells("A10:F10")
wc["A10"] = "RESULTATS (calcules automatiquement)"
wc["A10"].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
wc["A10"].fill = PatternFill("solid", fgColor=GREEN_MED)
wc["A10"].alignment = Alignment(horizontal="left")
wc.row_dimensions[10].height = 20

results = [
    ("Loyer de base / PDC / mois HT", '=IF(B5="Starter",89,IF(B5="Standard",119,IF(B5="Pro",139,IF(B5="Premium",169,0))))'),
    ("Loyer mensuel total HT (base)", "=B11*B6"),
    ("Supplement options (reparti sur duree)", "=IFERROR(B8/B7,0)"),
    ("Loyer mensuel total HT (options incluses)", "=B12+B13"),
    ("TVA Guyane 8,5 %", "=B14*0.085"),
    ("Loyer mensuel TTC", "=B14+B15"),
    ("", ""),
    ("Cout total LOA HT (loyers)", "=B14*B7"),
    ("Valeur residuelle option d'achat HT", '=IF(B5="Starter",72,IF(B5="Standard",160,IF(B5="Pro",274,IF(B5="Premium",495,0))))'),
    ("Cout total si option achat exercee HT", "=B18+B19"),
    ("", ""),
    ("Estimation travaux installation (bas)", 500),
    ("Estimation travaux installation (haut)", 4000),
    ("Investissement total client 72 mois (bas) HT", "=B18+B19+B22"),
    ("Investissement total client 72 mois (haut) HT", "=B18+B19+B23"),
]

for r, (label, val) in enumerate(results, 11):
    if not label:
        continue
    row_r = r
    wc[f"A{row_r}"] = label
    wc[f"A{row_r}"].font = Font(name="Arial", size=10)
    wc[f"A{row_r}"].alignment = Alignment(vertical="center")
    c = wc[f"B{row_r}"]
    c.value = val
    if str(val).startswith("="):
        c.number_format = '#,##0.00" EUR"'
        c.font = Font(name="Arial", size=11, bold=True, color="000000")
    else:
        c.font = Font(name="Arial", size=11, color="0000FF")
    c.fill = PatternFill("solid", fgColor=GREEN_LIGHT)
    wc.merge_cells(f"B{row_r}:F{row_r}")
    c.alignment = Alignment(horizontal="center", vertical="center")

# Highlight key rows
for key_row in [16, 18, 19, 20, 24, 25]:
    try:
        wc[f"A{key_row}"].font = Font(name="Arial", size=10, bold=True, color=GREEN_DARK)
        wc[f"B{key_row}"].font = Font(name="Arial", size=12, bold=True, color=GREEN_DARK)
        wc[f"B{key_row}"].fill = PatternFill("solid", fgColor="c8e6c9")
    except Exception:
        pass

# Section 3 : Tableau comparatif
wc.merge_cells("A28:F28")
wc["A28"] = "TABLEAU COMPARATIF DES 4 OFFRES LOA"
wc["A28"].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
wc["A28"].fill = PatternFill("solid", fgColor=GREEN_DARK)
wc["A28"].alignment = Alignment(horizontal="left")
wc.row_dimensions[28].height = 20

comp_headers = ["Critere", "Starter", "Standard", "Pro", "Premium"]
for col, hdr in enumerate(comp_headers, 1):
    c = wc.cell(row=29, column=col, value=hdr)
    bg = {"Starter": "fff9c4", "Standard": "c8e6c9", "Pro": "bbdefb", "Premium": "ce93d8"}.get(hdr, GREEN_DARK)
    fg = "FFFFFF" if hdr == "Critere" else "000000"
    c.fill = PatternFill("solid", fgColor=GREEN_DARK if hdr == "Critere" else bg)
    c.font = Font(name="Arial", size=10, bold=True, color=fg if hdr != "Critere" else "FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_thin()
    wc.row_dimensions[29].height = 18

comp_data = [
    ("Modele", "e-WallBox 1x7kW", "e-Smart 1x22kW", "e-Smart 2x22kW", "e-Premium 2x22kW"),
    ("Nombre de PDC", "1", "1", "2", "2"),
    ("Puissance totale", "7 kW", "22 kW", "44 kW", "44 kW"),
    ("Valeur materiel HT", "715 EUR", "1 599 EUR", "2 739 EUR", "4 948 EUR"),
    ("Loyer / PDC / mois HT", "89 EUR", "119 EUR", "139 EUR", "169 EUR"),
    ("Loyer total / mois HT", "89 EUR", "119 EUR", "278 EUR", "338 EUR"),
    ("Loyer total 72 mois HT", "6 408 EUR", "8 568 EUR", "20 016 EUR", "24 336 EUR"),
    ("Option d'achat HT", "72 EUR", "160 EUR", "274 EUR", "495 EUR"),
    ("Ecran 10 pouces (option)", "Non dispo", "+648 EUR", "+648 EUR", "+648 EUR"),
    ("TPE CB integre (option)", "Non dispo", "+837 EUR", "+837 EUR", "+837 EUR"),
    ("Design premium", "Non", "Non", "Non", "Oui"),
    ("Cible principale", "Bureaux, PME", "Commerces, banques", "Hotels, centres comm.", "CSG, aeroport, 4*"),
]

for r, row_data in enumerate(comp_data, 30):
    wc.row_dimensions[r].height = 15
    bg = ALT_ROW if r % 2 == 0 else "FFFFFF"
    for col, val in enumerate(row_data, 1):
        c = wc.cell(row=r, column=col, value=val)
        c.font = Font(name="Arial", size=9, bold=(col == 1))
        c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
        c.border = border_thin()
        c.fill = PatternFill("solid", fgColor=bg if col == 1 else bg)

# Tab color
wc.sheet_properties.tabColor = "2e7d32"
ws.sheet_properties.tabColor = "1a472a"

# Save
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"[OK] {OUT.name} genere")
print(f"     Sheet 1 : {len(TARGETS)} entreprises cibles")
print(f"     Sheet 2 : Calculateur LOA (4 offres + tableau comparatif)")
