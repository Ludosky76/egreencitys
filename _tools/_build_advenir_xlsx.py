"""Construit le tableur de chiffrage ADVENIR avec formules dynamiques."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(r"C:\projet\Egreencity\_dossiers\ADVENIR\2-chiffrage-advenir-bornes.xlsx")

# Couleurs charte EGREENCITY'S
GREEN_DARK  = "0A4800"
GREEN       = "33CC00"
GREEN_LIGHT = "E8F8E0"
BLUE_DARK   = "2B4DB5"
GOLD        = "D4AF37"
WHITE       = "FFFFFF"
GRAY        = "F4F4F4"

BLUE_FONT   = Font(name="Arial", color="0000FF")           # entrees hardcodees
BLACK_FONT  = Font(name="Arial", color="000000")           # formules
GREEN_FONT  = Font(name="Arial", color="008000")           # liens autres feuilles
BOLD_WHITE  = Font(name="Arial", bold=True, color=WHITE)
BOLD_DARK   = Font(name="Arial", bold=True, color=GREEN_DARK)

thin = Side(border_style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

EUR     = '_-* #,##0 €_-;[Red]-* #,##0 €_-;_-* "-" €_-;_-@_-'
EUR2    = '_-* #,##0.00 €_-;[Red]-* #,##0.00 €_-;_-* "-" €_-;_-@_-'
PCT     = '0.0%'
NUM     = '#,##0'

wb = Wb = Workbook()


# ================================================================
#  Feuille 1 : BAREME (parametrable)
# ================================================================
s1 = wb.active
s1.title = "Bareme ADVENIR"
s1.sheet_view.showGridLines = False

s1["A1"] = "BAREME ADVENIR — Outre-mer (à actualiser sur advenir.mobi)"
s1["A1"].font = Font(name="Arial", bold=True, size=14, color=GREEN_DARK)
s1.merge_cells("A1:F1")

headers = ["Cas d'usage", "Code", "Aide max / point (€)", "Taux max (% HT)",
           "Plafond installation (€)", "Notes"]
for i, h in enumerate(headers, 1):
    c = s1.cell(row=3, column=i, value=h)
    c.font = BOLD_WHITE
    c.fill = PatternFill("solid", fgColor=GREEN_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

# Lignes de bareme — valeurs typiques 2026 pour DOM (a verifier sur advenir.mobi)
rows = [
    ("Residentiel collectif (parking copro/social)",  "RES_COL", 960,  0.50, 1920,
     "Outre-mer : majoration possible de 10–20%. Verifier le bareme."),
    ("Salaries — point individuel",                    "SAL_IND", 960,  0.50, 1920,
     "Borne dediee a un salarie identifie."),
    ("Salaries — point partage",                       "SAL_PAR", 600,  0.50, 1200,
     "Plusieurs salaries se partagent la borne."),
    ("Flotte/visiteurs entreprise",                    "FLT_ENT", 1500, 0.50, 3000,
     "Parking employeur ouvert visiteurs."),
    ("Voirie publique — AC",                           "VOI_AC",  2200, 0.30, 7330,
     "Borne installee en voirie communale."),
    ("Voirie publique — DC (rapide)",                  "VOI_DC",  8100, 0.30, 27000,
     "Charge rapide DC ≥ 50 kW."),
    ("Parking ouvert public (prive ouvert) — AC",      "POP_AC",  1700, 0.20, 8500,
     "Centre commercial, hotel, etc."),
    ("Parking ouvert public — DC",                     "POP_DC",  9000, 0.20, 45000,
     "Charge rapide DC ≥ 50 kW."),
    ("Poids lourds (PL)",                              "PL_DC",   960000, 0.40, 2400000,
     "Bornes haute puissance ≥ 150 kW."),
]

for ridx, row in enumerate(rows, start=4):
    for cidx, val in enumerate(row, start=1):
        c = s1.cell(row=ridx, column=cidx, value=val)
        c.font = BLUE_FONT  # entrees parametrables -> bleu
        c.border = border
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if cidx in (3, 5):
            c.number_format = EUR
        elif cidx == 4:
            c.number_format = PCT

s1.column_dimensions["A"].width = 42
s1.column_dimensions["B"].width = 12
s1.column_dimensions["C"].width = 18
s1.column_dimensions["D"].width = 16
s1.column_dimensions["E"].width = 22
s1.column_dimensions["F"].width = 50

s1["A15"] = "Note : ces montants sont indicatifs et issus du bareme outre-mer 2026."
s1["A15"].font = Font(name="Arial", italic=True, color="888888", size=9)
s1["A16"] = "Verifier les valeurs en vigueur avant depot sur https://advenir.mobi"
s1["A16"].font = Font(name="Arial", italic=True, color="888888", size=9)
s1.merge_cells("A15:F15")
s1.merge_cells("A16:F16")


# ================================================================
#  Feuille 2 : CHIFFRAGE BORNES
# ================================================================
s2 = wb.create_sheet("Chiffrage")
s2.sheet_view.showGridLines = False

s2["A1"] = "EGREENCITY'S — Chiffrage ADVENIR par site"
s2["A1"].font = Font(name="Arial", bold=True, size=14, color=GREEN_DARK)
s2.merge_cells("A1:L1")

s2["A2"] = "Editer les colonnes en BLEU. Les colonnes en NOIR sont calculees automatiquement."
s2["A2"].font = Font(name="Arial", italic=True, color="666666", size=9)
s2.merge_cells("A2:L2")

cols = [
    ("#", 5),
    ("Commune", 18),
    ("Site / Adresse", 28),
    ("Type de borne", 22),
    ("Cas d'usage (code)", 16),
    ("Nb points", 9),
    ("Cout HT / point (€)", 16),
    ("Cout HT total (€)", 16),
    ("Aide unitaire bareme (€)", 17),
    ("Taux max (% HT)", 14),
    ("Aide ADVENIR calculee (€)", 19),
    ("Reste a charge HT (€)", 17),
]
for i, (h, w) in enumerate(cols, 1):
    c = s2.cell(row=4, column=i, value=h)
    c.font = BOLD_WHITE
    c.fill = PatternFill("solid", fgColor=GREEN_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    s2.column_dimensions[get_column_letter(i)].width = w
s2.row_dimensions[4].height = 32

# Donnees indicatives a remplir (BLEU)
sites = [
    (1,  "Cayenne",                 "Hotel de Ville — parking",         "Totem 5 departs (110 kVa)", "VOI_AC", 5, 11400),
    (2,  "Cayenne",                 "Marche central — voirie",          "Totem 3 departs (36 kVa)",  "VOI_AC", 3, 7100),
    (3,  "Cayenne",                 "CHAR — parking visiteurs",         "Pied double 2x22 kW",       "FLT_ENT", 4, 5670),
    (4,  "Kourou",                  "CSG — parking salaries",           "Pied double 2x22 kW",       "SAL_IND", 6, 5670),
    (5,  "Kourou",                  "Front de mer — voirie",            "Totem 3 departs (36 kVa)",  "VOI_AC", 2, 7100),
    (6,  "Saint-Laurent-du-Maroni", "Centre administratif",             "Totem 3 departs (36 kVa)",  "VOI_AC", 3, 7100),
    (7,  "Saint-Laurent-du-Maroni", "Hopital — visiteurs",              "Pied solo 1x22 kW",         "POP_AC", 3, 4200),
    (8,  "Matoury",                 "Centre commercial Family Plaza",   "Pied double 2x22 kW",       "POP_AC", 4, 5670),
    (9,  "Macouria",                "RN1 — aire de service",            "Borne rapide DC 120 kW",    "VOI_DC", 2, 54500),
    (10, "Remire-Montjoly",         "Plage — parking communal",         "Pied double 2x22 kW",       "VOI_AC", 4, 5670),
    (11, "Mana",                    "Mairie",                           "Pied solo 1x22 kW",         "VOI_AC", 2, 4200),
    (12, "Iracoubo",                "Mairie",                           "Pied solo 1x22 kW",         "VOI_AC", 2, 4200),
    (13, "Roura",                   "Mairie",                           "Pied solo 1x22 kW",         "VOI_AC", 2, 4200),
    (14, "Sinnamary",               "Bord de mer",                      "Totem 3 departs (36 kVa)",  "VOI_AC", 3, 7100),
    (15, "Macouria",                "Residence collective La Rougerie", "WallBox murale 22 kW",      "RES_COL", 6, 1620),
]

start = 5
for i, s in enumerate(sites):
    r = start + i
    # Colonnes saisie (BLEU)
    s2.cell(row=r, column=1, value=s[0]).font = BLUE_FONT
    s2.cell(row=r, column=2, value=s[1]).font = BLUE_FONT
    s2.cell(row=r, column=3, value=s[2]).font = BLUE_FONT
    s2.cell(row=r, column=4, value=s[3]).font = BLUE_FONT
    s2.cell(row=r, column=5, value=s[4]).font = BLUE_FONT
    s2.cell(row=r, column=6, value=s[5]).font = BLUE_FONT
    s2.cell(row=r, column=7, value=s[6]).font = BLUE_FONT

    # Cout HT total = nb x cout/point
    s2.cell(row=r, column=8, value=f"=F{r}*G{r}").font = BLACK_FONT

    # Aide unitaire bareme = VLOOKUP par code
    s2.cell(row=r, column=9,
            value=f"=IFERROR(VLOOKUP(E{r},'Bareme ADVENIR'!$B$4:$E$13,2,FALSE),0)").font = GREEN_FONT
    # Taux max = VLOOKUP
    s2.cell(row=r, column=10,
            value=f"=IFERROR(VLOOKUP(E{r},'Bareme ADVENIR'!$B$4:$E$13,3,FALSE),0)").font = GREEN_FONT
    # Aide calculee = MIN(aide_unitaire * nb ; cout_total * taux)
    s2.cell(row=r, column=11,
            value=f"=MIN(I{r}*F{r},H{r}*J{r})").font = BLACK_FONT
    # Reste a charge = cout total - aide
    s2.cell(row=r, column=12, value=f"=H{r}-K{r}").font = BLACK_FONT

    for cidx in range(1, 13):
        s2.cell(row=r, column=cidx).border = border
        s2.cell(row=r, column=cidx).alignment = Alignment(vertical="center", wrap_text=True)

    for cidx in (7, 8, 9, 11, 12):
        s2.cell(row=r, column=cidx).number_format = EUR
    s2.cell(row=r, column=10).number_format = PCT

    if i % 2 == 1:
        for cidx in range(1, 13):
            s2.cell(row=r, column=cidx).fill = PatternFill("solid", fgColor=GREEN_LIGHT)

# Ligne de totaux
total_row = start + len(sites) + 1
s2.cell(row=total_row, column=1, value="TOTAL").font = BOLD_DARK
s2.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor=GOLD)
s2.merge_cells(start_row=total_row, start_column=1,
               end_row=total_row, end_column=5)
s2.cell(row=total_row, column=6, value=f"=SUM(F{start}:F{start+len(sites)-1})").font = BOLD_DARK
s2.cell(row=total_row, column=8, value=f"=SUM(H{start}:H{start+len(sites)-1})").font = BOLD_DARK
s2.cell(row=total_row, column=11, value=f"=SUM(K{start}:K{start+len(sites)-1})").font = BOLD_DARK
s2.cell(row=total_row, column=12, value=f"=SUM(L{start}:L{start+len(sites)-1})").font = BOLD_DARK
for cidx in (6, 8, 11, 12):
    cell = s2.cell(row=total_row, column=cidx)
    cell.fill = PatternFill("solid", fgColor=GOLD)
    cell.border = border
    cell.font = Font(name="Arial", bold=True, color="000000")
    if cidx != 6:
        cell.number_format = EUR
    else:
        cell.number_format = NUM

# Indicateurs cles
ki = total_row + 3
s2.cell(row=ki, column=1, value="INDICATEURS CLES").font = Font(name="Arial", bold=True, size=12, color=GREEN_DARK)
s2.merge_cells(start_row=ki, start_column=1, end_row=ki, end_column=5)

ki += 2
s2.cell(row=ki, column=1, value="Nb total de points de charge").font = BOLD_DARK
s2.cell(row=ki, column=3, value=f"=F{total_row}").font = BLACK_FONT
s2.cell(row=ki, column=3).number_format = NUM

ki += 1
s2.cell(row=ki, column=1, value="Investissement total HT").font = BOLD_DARK
s2.cell(row=ki, column=3, value=f"=H{total_row}").font = BLACK_FONT
s2.cell(row=ki, column=3).number_format = EUR

ki += 1
s2.cell(row=ki, column=1, value="Aide ADVENIR sollicitee").font = BOLD_DARK
s2.cell(row=ki, column=3, value=f"=K{total_row}").font = BLACK_FONT
s2.cell(row=ki, column=3).number_format = EUR
s2.cell(row=ki, column=3).fill = PatternFill("solid", fgColor="FFFF00")  # surbrillance

ki += 1
s2.cell(row=ki, column=1, value="Taux moyen d'aide / investissement").font = BOLD_DARK
s2.cell(row=ki, column=3, value=f"=K{total_row}/H{total_row}").font = BLACK_FONT
s2.cell(row=ki, column=3).number_format = PCT

ki += 1
s2.cell(row=ki, column=1, value="Reste a charge HT (autofinancement)").font = BOLD_DARK
s2.cell(row=ki, column=3, value=f"=L{total_row}").font = BLACK_FONT
s2.cell(row=ki, column=3).number_format = EUR


# ================================================================
#  Feuille 3 : PLAN DE FINANCEMENT
# ================================================================
s3 = wb.create_sheet("Plan financement")
s3.sheet_view.showGridLines = False

s3["A1"] = "Plan de financement previsionnel — Phase 1"
s3["A1"].font = Font(name="Arial", bold=True, size=14, color=GREEN_DARK)
s3.merge_cells("A1:D1")

headers = ["Source", "Type", "Montant HT (€)", "Part (% du total)"]
for i, h in enumerate(headers, 1):
    c = s3.cell(row=3, column=i, value=h)
    c.font = BOLD_WHITE
    c.fill = PatternFill("solid", fgColor=GREEN_DARK)
    c.alignment = Alignment(horizontal="center")
    c.border = border

financements = [
    ("ADVENIR (subvention)",      "Aide publique",   f"=Chiffrage!K{total_row}", "GREEN"),
    ("Apport associe(s)",         "Fonds propres",   30000, "BLUE"),
    ("Emprunt bancaire",          "Dette",           80000, "BLUE"),
    ("Subvention CTG / FEDER",    "Aide regionale",  20000, "BLUE"),
    ("Autofinancement (CAF)",     "Tresorerie",      0, "BLUE"),
]

s3.column_dimensions["A"].width = 36
s3.column_dimensions["B"].width = 18
s3.column_dimensions["C"].width = 20
s3.column_dimensions["D"].width = 16

for i, (src, typ, mt, src_color) in enumerate(financements, start=4):
    s3.cell(row=i, column=1, value=src).font = BOLD_DARK
    s3.cell(row=i, column=2, value=typ).font = BLUE_FONT
    cm = s3.cell(row=i, column=3, value=mt)
    cm.font = GREEN_FONT if src_color == "GREEN" else BLUE_FONT
    cm.number_format = EUR
    s3.cell(row=i, column=4, value=f"=C{i}/SUM($C$4:$C$8)").font = BLACK_FONT
    s3.cell(row=i, column=4).number_format = PCT
    for cidx in range(1, 5):
        s3.cell(row=i, column=cidx).border = border

s3.cell(row=10, column=1, value="TOTAL").font = BOLD_WHITE
s3.cell(row=10, column=1).fill = PatternFill("solid", fgColor=GREEN_DARK)
s3.cell(row=10, column=3, value="=SUM(C4:C8)").font = BOLD_WHITE
s3.cell(row=10, column=3).fill = PatternFill("solid", fgColor=GREEN_DARK)
s3.cell(row=10, column=3).number_format = EUR
s3.cell(row=10, column=4, value="=SUM(D4:D8)").font = BOLD_WHITE
s3.cell(row=10, column=4).fill = PatternFill("solid", fgColor=GREEN_DARK)
s3.cell(row=10, column=4).number_format = PCT
for cidx in range(1, 5):
    s3.cell(row=10, column=cidx).border = border

# Note coherence
s3["A12"] = "Le total doit egaler l'investissement HT calcule en feuille 'Chiffrage'."
s3["A12"].font = Font(name="Arial", italic=True, color="666666", size=10)
s3["A13"] = "Chiffrage feuille 2 — Investissement HT total :"
s3["A13"].font = BOLD_DARK
s3["C13"] = f"=Chiffrage!H{total_row}"
s3["C13"].font = GREEN_FONT
s3["C13"].number_format = EUR
s3["A14"] = "Plan financement total :"
s3["A14"].font = BOLD_DARK
s3["C14"] = "=C10"
s3["C14"].font = BLACK_FONT
s3["C14"].number_format = EUR
s3["A15"] = "Ecart (doit etre 0) :"
s3["A15"].font = BOLD_DARK
s3["C15"] = "=C13-C14"
s3["C15"].font = BLACK_FONT
s3["C15"].number_format = EUR
s3["C15"].fill = PatternFill("solid", fgColor="FFFF00")


# ================================================================
#  Sauvegarde
# ================================================================
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"OK -> {OUT}")
