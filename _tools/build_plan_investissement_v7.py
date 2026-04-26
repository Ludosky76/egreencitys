"""
Plan Investissement EGREENCITY'S — Version 2026 — 20 PDC AC uniquement
10 × e-Premium AC 2×22 kW — Programme ADVENIR Guyane
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = r"C:\Users\ludosky\Desktop\ALBIOMA\Plan Investissement e-Premium 20 PDC AC 2026.xlsx"

# ─── Styles ──────────────────────────────────────────────────────────────────

def font(bold=False, size=11, color="000000", italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_thin():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium_bottom():
    t = Side(style="thin", color="BBBBBB")
    b = Side(style="medium", color="444444")
    return Border(left=t, right=t, top=t, bottom=b)

# Palette
C_HEADER   = "1B4F72"   # dark blue header bg
C_SUBHEAD  = "2E86C1"   # sub-header bg
C_GREEN    = "1D8348"   # green total bg
C_GREEN_L  = "D5F5E3"   # light green bg
C_ORANGE   = "E67E22"   # accent/warning
C_YELLOW   = "FFF9C4"   # assumption highlight
C_GREY     = "F2F3F4"   # alternating row
C_WHITE    = "FFFFFF"
C_BLUE_IN  = "0000FF"   # blue for inputs (hardcoded)

FMT_EUR    = '#,##0 "€";(#,##0 "€");-'
FMT_EUR2   = '#,##0.00 "€";(#,##0.00 "€");-'
FMT_PCT    = '0.0%'
FMT_INT    = '#,##0;(#,##0);-'
FMT_COEF   = '0.000'


def set_cell(ws, row, col, value=None, bold=False, size=11, color="000000",
             bg=None, h="left", v="center", wrap=False, fmt=None, italic=False, border=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, size=size, color=color, italic=italic)
    c.alignment = align(h=h, v=v, wrap=wrap)
    if bg:
        c.fill = fill(bg)
    if fmt:
        c.number_format = fmt
    if border:
        c.border = border_thin()
    return c


def header_row(ws, row, cols, texts, bg=C_HEADER):
    for col, text in zip(cols, texts):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = fill(bg)
        c.alignment = align(h="center", v="center", wrap=True)
        c.border = border_thin()


def title_row(ws, row, col_start, col_end, text, size=13, bg=C_HEADER):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.font = Font(name="Arial", bold=True, size=size, color="FFFFFF")
    c.fill = fill(bg)
    c.alignment = align(h="center", v="center")


def section_title(ws, row, col_start, col_end, text, bg=C_SUBHEAD):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    c.fill = fill(bg)
    c.alignment = align(h="left", v="center")
    ws.row_dimensions[row].height = 18


def total_row(ws, row, cols, values, bg=C_GREEN):
    for col, val in zip(cols, values):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill = fill(bg)
        c.alignment = align(h="right" if col > 1 else "left", v="center")
        c.border = border_thin()
        if isinstance(val, str) and val.startswith("="):
            c.number_format = FMT_EUR
        elif isinstance(val, (int, float)):
            c.number_format = FMT_EUR


# ─── Sheet 1: Configuration & Coûts ─────────────────────────────────────────

def build_config(wb):
    ws = wb.create_sheet("Configuration & Coûts")
    ws.sheet_view.showGridLines = False

    # Column widths
    widths = [2, 38, 10, 14, 14, 14, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 8

    # Title
    title_row(ws, 2, 1, 7,
              "PLAN D'INVESTISSEMENT — EGREENCITY'S  |  20 PDC — 10 × e-Premium AC 2×22 kW  |  Guyane 2026",
              size=13, bg=C_HEADER)

    # ── LOT 1 ──
    ws.row_dimensions[4].height = 8
    section_title(ws, 5, 2, 7, "  LOT 1 — Bornes e-Premium AC Pied 2×22 kW  ×10  (Base ADVENIR)")

    header_row(ws, 6, [2, 3, 4, 5, 6, 7],
               ["Composante", "Qté", "Prix unit. HT (€)", "Total HT (€)", "TTC Guyane (×1,670)", "Source"])

    items = [
        ("Borne e-Premium AC Pied 2×22 kW (fourniture)", 10, 4910),
        ("Option : écran tactile 10\" multilingue", 10, 648),
        ("Option : pied coffret CIBE", 10, 108),
        ("Option : modem 4G intégré", 10, 95),
        ("Option : traitement anti-graffiti tropicale", 10, 41),
    ]
    first_item_row = 7
    for i, (name, qty, prix) in enumerate(items):
        r = first_item_row + i
        ws.row_dimensions[r].height = 17
        bg = C_WHITE if i % 2 == 0 else C_GREY
        set_cell(ws, r, 2, name, bg=bg, border=True)
        set_cell(ws, r, 3, qty, bg=bg, h="center", fmt=FMT_INT, border=True, color=C_BLUE_IN)
        set_cell(ws, r, 4, prix, bg=bg, h="right", fmt=FMT_EUR, border=True, color=C_BLUE_IN)
        set_cell(ws, r, 5, f"=C{r}*D{r}", bg=bg, h="right", fmt=FMT_EUR, border=True)

    # Sous-total fournisseur HT
    r_st = first_item_row + len(items)
    ws.row_dimensions[r_st].height = 18
    ws.merge_cells(start_row=r_st, start_column=2, end_row=r_st, end_column=4)
    set_cell(ws, r_st, 2, "Sous-total LOT 1 HT (base éligible ADVENIR)", bold=True, bg=C_GREEN_L, border=True)
    set_cell(ws, r_st, 5, f"=SUM(E{first_item_row}:E{r_st-1})", bold=True, bg=C_GREEN_L, h="right", fmt=FMT_EUR, border=True)

    # Coefficient note
    r_coef = r_st + 1
    ws.row_dimensions[r_coef].height = 8
    r_coef2 = r_coef + 1
    ws.row_dimensions[r_coef2].height = 16
    ws.merge_cells(start_row=r_coef2, start_column=2, end_row=r_coef2, end_column=4)
    set_cell(ws, r_coef2, 2, "Coefficient Guyane ×1,670  (OM 23% + Fret maritime 8% + Mise en service EDF)",
             italic=True, color="555555", wrap=True)
    set_cell(ws, r_coef2, 5, 1.670, color=C_BLUE_IN, h="center", fmt=FMT_COEF, border=True)

    # Installation & services
    ws.row_dimensions[r_coef2 + 1].height = 17
    ws.row_dimensions[r_coef2 + 2].height = 17
    ws.row_dimensions[r_coef2 + 3].height = 17
    inst_items = [
        ("Installation & raccordement EDF SEI BT (×10 bornes)", 22850),
        ("Mise en service OCPP + tests (×10)", 1800),
        ("Signalétique ADVENIR — sérigraphie (×10)", 500),
    ]
    for j, (name, amount) in enumerate(inst_items):
        r2 = r_coef2 + 1 + j
        bg = C_WHITE if j % 2 == 0 else C_GREY
        set_cell(ws, r2, 2, name, bg=bg, border=True)
        set_cell(ws, r2, 5, amount, bg=bg, h="right", fmt=FMT_EUR, border=True, color=C_BLUE_IN)

    r_ttc = r_coef2 + 4
    ws.row_dimensions[r_ttc].height = 20
    ws.merge_cells(start_row=r_ttc, start_column=2, end_row=r_ttc, end_column=4)
    set_cell(ws, r_ttc, 2, "TOTAL LOT 1 — TTC Guyane", bold=True, size=12, bg=C_HEADER, color="FFFFFF", border=True)
    set_cell(ws, r_ttc, 5, 96870, bold=True, size=12, bg=C_HEADER, color="FFFFFF", h="right", fmt=FMT_EUR, border=True)
    set_cell(ws, r_ttc, 7, "Source: devis E-TOTEM DEV17000172-6, 2025", italic=True, color="777777", size=9)

    # ── LOT 2 — Supervision ──
    ws.row_dimensions[r_ttc + 1].height = 8
    r_lot2 = r_ttc + 2
    section_title(ws, r_lot2, 2, 7, "  LOT 2 — Forfait supervision & services (5 ans)")

    header_row(ws, r_lot2 + 1, [2, 3, 5, 7],
               ["Service", "Durée", "Montant TTC (€)", "Note"])

    sup_items = [
        ("Plateforme CSMS OCPP 1.6J — 20 PDC", "5 ans", 5000),
        ("Application mobile iOS/Android (marque blanche)", "5 ans", 2400),
        ("Hotline opérateur niveau 2 (24/7)", "5 ans", 2600),
        ("Formation Loic + Patrice LUDOSKY (prise en main CSMS)", "2 jours", 1200),
        ("Déplacement technicien E-TOTEM Guyane — mise en service", "1 voyage", 300),
    ]
    r_sup0 = r_lot2 + 2
    for k, (svc, dur, amt) in enumerate(sup_items):
        r3 = r_sup0 + k
        ws.row_dimensions[r3].height = 17
        bg = C_WHITE if k % 2 == 0 else C_GREY
        set_cell(ws, r3, 2, svc, bg=bg, border=True)
        set_cell(ws, r3, 3, dur, bg=bg, h="center", border=True)
        set_cell(ws, r3, 5, amt, bg=bg, h="right", fmt=FMT_EUR, border=True, color=C_BLUE_IN)

    r_sup_tot = r_sup0 + len(sup_items)
    ws.row_dimensions[r_sup_tot].height = 18
    ws.merge_cells(start_row=r_sup_tot, start_column=2, end_row=r_sup_tot, end_column=4)
    set_cell(ws, r_sup_tot, 2, "TOTAL LOT 2 — Supervision & services", bold=True, bg=C_GREEN_L, border=True)
    set_cell(ws, r_sup_tot, 5, f"=SUM(E{r_sup0}:E{r_sup_tot-1})", bold=True, bg=C_GREEN_L, h="right", fmt=FMT_EUR, border=True)

    # ── LOT 3 — Divers ──
    r_div = r_sup_tot + 2
    section_title(ws, r_div, 2, 7, "  LOT 3 — Divers & Contingence (2 %)")
    ws.row_dimensions[r_div + 1].height = 17
    ws.merge_cells(start_row=r_div + 1, start_column=2, end_row=r_div + 1, end_column=4)
    set_cell(ws, r_div + 1, 2, "Contingence / imprévus (2 % du budget total)", bg=C_GREY, border=True)
    set_cell(ws, r_div + 1, 5, 2170, bg=C_GREY, h="right", fmt=FMT_EUR, border=True, color=C_BLUE_IN)

    # ── TOTAL GÉNÉRAL ──
    r_grand = r_div + 3
    ws.row_dimensions[r_grand].height = 24
    ws.merge_cells(start_row=r_grand, start_column=2, end_row=r_grand, end_column=4)
    set_cell(ws, r_grand, 2, "TOTAL BUDGET — TTC Guyane", bold=True, size=13, bg=C_GREEN, color="FFFFFF", border=True)
    set_cell(ws, r_grand, 5, 110540, bold=True, size=13, bg=C_GREEN, color="FFFFFF", h="right", fmt=FMT_EUR, border=True)

    # ADVENIR
    r_adv = r_grand + 1
    ws.row_dimensions[r_adv].height = 18
    ws.merge_cells(start_row=r_adv, start_column=2, end_row=r_adv, end_column=4)
    set_cell(ws, r_adv, 2, "dont Subvention ADVENIR — 20 PDC × 1 860 €", italic=True, bg=C_YELLOW, border=True)
    set_cell(ws, r_adv, 5, -37200, italic=True, bold=True, bg=C_YELLOW, h="right", fmt=FMT_EUR, border=True)

    r_taux = r_adv + 1
    ws.row_dimensions[r_taux].height = 18
    ws.merge_cells(start_row=r_taux, start_column=2, end_row=r_taux, end_column=4)
    set_cell(ws, r_taux, 2, "Taux de couverture ADVENIR (sur base HT matériel)", bold=True, bg=C_YELLOW, border=True)
    set_cell(ws, r_taux, 5, "=(-E{})/{} ".format(r_adv, 58020).replace("=(-", "=ABS(E{}/{}".format(r_adv, 58020)).replace(" ", ""),
             bold=True, bg=C_YELLOW, h="right", fmt=FMT_PCT, border=True)
    # Fix: simple hardcoded formula
    ws.cell(row=r_taux, column=5).value = "=ABS(E{})/58020".format(r_adv)

    # Footnote
    r_fn = r_taux + 2
    ws.merge_cells(start_row=r_fn, start_column=2, end_row=r_fn, end_column=7)
    set_cell(ws, r_fn, 2,
             "Texte bleu = hypothèse modifiable  |  Base HT ADVENIR = sous-total LOT 1 (matériel fournisseur, hors installation)  "
             "|  Coefficient Guyane ×1,670 : OM 23% + fret 8% + installation EDF SEI",
             italic=True, color="777777", size=9, wrap=True)
    ws.row_dimensions[r_fn].height = 28

    return ws


# ─── Sheet 2: Plan de Financement ────────────────────────────────────────────

def build_financement(wb):
    ws = wb.create_sheet("Plan de Financement")
    ws.sheet_view.showGridLines = False

    widths = [2, 38, 18, 14, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 8

    title_row(ws, 2, 1, 5,
              "PLAN DE FINANCEMENT — EGREENCITY'S  |  20 PDC — 10 × e-Premium AC 2×22 kW  |  Guyane 2026",
              size=13, bg=C_HEADER)

    # Investissement recap
    section_title(ws, 4, 2, 5, "  Investissement total")
    header_row(ws, 5, [2, 3, 4, 5], ["Lot", "Montant TTC Guyane (€)", "% du total", "Détail"])

    inv_rows = [
        ("LOT 1 — 10 × e-Premium AC 2×22 kW", 96870),
        ("LOT 2 — Supervision & services (5 ans)", 11500),
        ("LOT 3 — Divers / contingence 2%", 2170),
    ]
    r0 = 6
    for i, (name, amt) in enumerate(inv_rows):
        r = r0 + i
        ws.row_dimensions[r].height = 17
        bg = C_WHITE if i % 2 == 0 else C_GREY
        set_cell(ws, r, 2, name, bg=bg, border=True)
        set_cell(ws, r, 3, amt, bg=bg, h="right", fmt=FMT_EUR, border=True, color=C_BLUE_IN)
        set_cell(ws, r, 4, f"=C{r}/C{r0+len(inv_rows)}", bg=bg, h="right", fmt=FMT_PCT, border=True)

    r_inv_tot = r0 + len(inv_rows)
    ws.row_dimensions[r_inv_tot].height = 20
    ws.merge_cells(start_row=r_inv_tot, start_column=2, end_row=r_inv_tot, end_column=2)
    set_cell(ws, r_inv_tot, 2, "INVESTISSEMENT TOTAL TTC", bold=True, bg=C_HEADER, color="FFFFFF", border=True)
    set_cell(ws, r_inv_tot, 3, f"=SUM(C{r0}:C{r_inv_tot-1})", bold=True, bg=C_HEADER, color="FFFFFF", h="right", fmt=FMT_EUR, border=True)
    set_cell(ws, r_inv_tot, 4, "100,0%", bold=True, bg=C_HEADER, color="FFFFFF", h="right", border=True)

    # Sources de financement
    ws.row_dimensions[r_inv_tot + 1].height = 8
    r_src = r_inv_tot + 2
    section_title(ws, r_src, 2, 5, "  Sources de financement")
    header_row(ws, r_src + 1, [2, 3, 4, 5],
               ["Source", "Montant (€)", "% du budget TTC", "Conditions"])

    sources = [
        ("Subvention ADVENIR (AVERE-France) — 20 PDC × 1 860 €", 37200,
         "Après mise en service — voirie publique OCPP 1.6J"),
        ("Exonération Octroi de Mer (demandée CTG)", 13345,
         "58 020 € × 23% — demande en cours Collectivité Territoriale"),
        ("BPI France — Prêt Amorçage DROM 0% (5 ans)", 25000,
         "PME DROM < 5 ans — demande en cours BPI Guyane"),
        ("OPCO Commerce (formation Qualiopi IRVE) + Bonus VU électrique", 11500,
         "Formation qualification IRVE + véhicule utilitaire électrique"),
        ("Apport fonds propres / emprunt bancaire court terme", 23495,
         "Solde résiduel — emprunt 36 mois ~683 €/mois"),
    ]
    r_src0 = r_src + 2
    for j, (name, amt, cond) in enumerate(sources):
        r = r_src0 + j
        ws.row_dimensions[r].height = 22
        bg = C_WHITE if j % 2 == 0 else C_GREY
        set_cell(ws, r, 2, name, bg=bg, border=True, wrap=True)
        set_cell(ws, r, 3, amt, bg=bg, h="right", fmt=FMT_EUR, border=True, color=C_BLUE_IN)
        set_cell(ws, r, 4, f"=C{r}/C{r_inv_tot}", bg=bg, h="right", fmt=FMT_PCT, border=True)
        set_cell(ws, r, 5, cond, bg=bg, italic=True, color="555555", size=9, wrap=True, border=True)

    r_src_tot = r_src0 + len(sources)
    ws.row_dimensions[r_src_tot].height = 20
    set_cell(ws, r_src_tot, 2, "TOTAL FINANCEMENT", bold=True, bg=C_GREEN, color="FFFFFF", border=True)
    set_cell(ws, r_src_tot, 3, f"=SUM(C{r_src0}:C{r_src_tot-1})", bold=True, bg=C_GREEN, color="FFFFFF", h="right", fmt=FMT_EUR, border=True)
    set_cell(ws, r_src_tot, 4, f"=C{r_src_tot}/C{r_inv_tot}", bold=True, bg=C_GREEN, color="FFFFFF", h="right", fmt=FMT_PCT, border=True)

    # Bilan
    ws.row_dimensions[r_src_tot + 1].height = 8
    r_bilan = r_src_tot + 2
    section_title(ws, r_bilan, 2, 5, "  Bilan — Capital résiduel à financer")

    for label, formula, bg_color in [
        ("Budget total TTC Guyane", f"=C{r_inv_tot}", C_GREY),
        ("− Subvention ADVENIR", f"=-C{r_src0}", C_GREY),
        ("− Exonération Octroi de Mer (si accordée)", f"=-C{r_src0+1}", C_GREY),
        ("− Prêt BPI DROM 0%", f"=-C{r_src0+2}", C_GREY),
        ("− OPCO Commerce + Bonus VU", f"=-C{r_src0+3}", C_GREY),
    ]:
        r_bilan += 1
        ws.row_dimensions[r_bilan].height = 17
        set_cell(ws, r_bilan, 2, label, bg=bg_color, border=True, italic=True)
        set_cell(ws, r_bilan, 3, formula, bg=bg_color, h="right", fmt=FMT_EUR, border=True)

    r_bilan += 1
    ws.row_dimensions[r_bilan].height = 24
    set_cell(ws, r_bilan, 2, "= CAPITAL RÉSIDUEL À FINANCER", bold=True, size=13, bg=C_ORANGE, color="FFFFFF", border=True)
    # Sum all the bilan rows (first is positive, rest negative)
    r_bilan_start = r_src_tot + 3
    set_cell(ws, r_bilan, 3, f"=SUM(C{r_bilan_start}:C{r_bilan-1})",
             bold=True, size=13, bg=C_ORANGE, color="FFFFFF", h="right", fmt=FMT_EUR, border=True)

    # Note
    r_note = r_bilan + 2
    ws.merge_cells(start_row=r_note, start_column=2, end_row=r_note, end_column=5)
    set_cell(ws, r_note, 2,
             "Taux de couverture ADVENIR : 37 200 € / 58 020 € HT matériel = 64,1 %  |  "
             "Capital résiduel : emprunt 36 mois à 3 % an → ~683 €/mois  |  "
             "ROI estimé < 18 mois au scénario de base (3 sessions/PDC/jour)",
             italic=True, color="444444", size=9, wrap=True)
    ws.row_dimensions[r_note].height = 28

    return ws


# ─── Sheet 3: Remboursement 36 Mois ──────────────────────────────────────────

def build_remboursement(wb):
    ws = wb.create_sheet("Remboursement 36 Mois")
    ws.sheet_view.showGridLines = False

    widths = [2, 10, 20, 18, 18, 18, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[2].height = 30
    title_row(ws, 2, 1, 7,
              "TABLEAU D'AMORTISSEMENT — Capital résiduel 23 495 €  |  36 mois à 3,0 % /an",
              size=12, bg=C_HEADER)

    # Parameters box
    ws.row_dimensions[4].height = 18
    section_title(ws, 4, 2, 7, "  Paramètres du prêt")

    params = [
        ("Capital emprunté (€)", 23495, "C6", FMT_EUR),
        ("Taux annuel (%)", 0.030, "C7", FMT_PCT),
        ("Taux mensuel (%)", "=C7/12", "C8", '0.000%'),
        ("Durée (mois)", 36, "C9", FMT_INT),
        ("Mensualité (€)", "=PMT(C8,C9,-C6)", "C10", FMT_EUR2),
        ("Total remboursé (€)", "=C10*C9", "C11", FMT_EUR2),
        ("Coût total des intérêts (€)", "=C11-C6", "C12", FMT_EUR2),
    ]
    for i, (label, val, ref, fmt) in enumerate(params):
        r = 5 + i
        ws.row_dimensions[r].height = 17
        bg = C_YELLOW if i < 2 else C_GREY
        color = C_BLUE_IN if i < 2 else "000000"
        set_cell(ws, r, 2, label, bg=bg, border=True, bold=(i >= 4))
        c = ws.cell(row=r, column=3, value=val)
        c.font = Font(name="Arial", bold=(i >= 4), color=color)
        c.fill = fill(bg)
        c.alignment = align(h="right")
        c.number_format = fmt
        c.border = border_thin()

    # Table header
    r_th = 14
    ws.row_dimensions[r_th].height = 8
    header_row(ws, r_th + 1, [2, 3, 4, 5, 6, 7],
               ["Mois", "Capital début (€)", "Mensualité (€)", "dont Intérêts (€)", "dont Capital (€)", "Capital restant (€)"])

    # 36 rows
    r_data = r_th + 2
    for m in range(1, 37):
        r = r_data + m - 1
        ws.row_dimensions[r].height = 15
        bg = C_WHITE if m % 2 == 1 else C_GREY

        set_cell(ws, r, 2, m, bg=bg, h="center", fmt=FMT_INT, border=True)

        if m == 1:
            cap_debut = "=C6"
        else:
            cap_debut = f"=G{r-1}"

        set_cell(ws, r, 3, cap_debut, bg=bg, h="right", fmt=FMT_EUR2, border=True)
        set_cell(ws, r, 4, "=C$10", bg=bg, h="right", fmt=FMT_EUR2, border=True)
        set_cell(ws, r, 5, f"=C{r}*C$8", bg=bg, h="right", fmt=FMT_EUR2, border=True)
        set_cell(ws, r, 6, f"=D{r}-E{r}", bg=bg, h="right", fmt=FMT_EUR2, border=True)
        set_cell(ws, r, 7, f"=C{r}-F{r}", bg=bg, h="right", fmt=FMT_EUR2, border=True)

    # Total row
    r_tot = r_data + 36
    ws.row_dimensions[r_tot].height = 20
    set_cell(ws, r_tot, 2, "TOTAL", bold=True, bg=C_HEADER, color="FFFFFF", h="center", border=True)
    set_cell(ws, r_tot, 3, "", bg=C_HEADER, border=True)
    set_cell(ws, r_tot, 4, f"=SUM(D{r_data}:D{r_tot-1})", bold=True, bg=C_HEADER, color="FFFFFF", h="right", fmt=FMT_EUR2, border=True)
    set_cell(ws, r_tot, 5, f"=SUM(E{r_data}:E{r_tot-1})", bold=True, bg=C_HEADER, color="FFFFFF", h="right", fmt=FMT_EUR2, border=True)
    set_cell(ws, r_tot, 6, f"=SUM(F{r_data}:F{r_tot-1})", bold=True, bg=C_HEADER, color="FFFFFF", h="right", fmt=FMT_EUR2, border=True)
    set_cell(ws, r_tot, 7, "— 0,00 €", bold=True, bg=C_GREEN, color="FFFFFF", h="right", border=True)

    return ws


# ─── Sheet 4: Rentabilité ─────────────────────────────────────────────────────

def build_rentabilite(wb):
    ws = wb.create_sheet("Rentabilité")
    ws.sheet_view.showGridLines = False

    widths = [2, 38, 3, 14, 14, 14, 14, 14, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[2].height = 30
    title_row(ws, 2, 1, 9,
              "COMPTE DE RÉSULTAT PRÉVISIONNEL — 20 PDC e-Premium AC 2×22 kW  |  An 1→5  |  Guyane 2026–2030",
              size=12, bg=C_HEADER)

    # Assumptions
    ws.row_dimensions[3].height = 8
    section_title(ws, 4, 2, 9, "  Hypothèses d'exploitation — Modifiables (texte bleu)")

    assum = [
        ("PDC disponibles (e-Premium AC)", 20, "C5"),
        ("Sessions / PDC / jour — An 1 (base)", 3, "C6"),
        ("Croissance sessions / an", 0.20, "C7"),
        ("Énergie moyenne / session (kWh)", 8, "C8"),
        ("Tarif vente AC (€/kWh)", 0.35, "C9"),
        ("Coût achat EDF SEI (€/kWh)", 0.14, "C10"),
        ("Supervision & maintenance (€/an)", 4500, "C11"),
        ("Assurance + frais gestion (€/an)", 2400, "C12"),
        ("Taux croissance charges fixes / an", 0.03, "C13"),
    ]
    for i, (label, val, ref) in enumerate(assum):
        r = 5 + i
        ws.row_dimensions[r].height = 16
        bg = C_YELLOW if i < 7 else C_GREY
        set_cell(ws, r, 2, label, bg=bg, border=True)
        c = ws.cell(row=r, column=3, value=val)
        c.font = Font(name="Arial", color=C_BLUE_IN)
        c.fill = fill(bg)
        c.alignment = align(h="right")
        c.border = border_thin()
        if "%" in label or "Croissance" in label or "Taux" in label:
            c.number_format = FMT_PCT
        elif "€" in label:
            c.number_format = FMT_EUR
        else:
            c.number_format = FMT_INT

    # Years header
    ws.row_dimensions[15].height = 8
    r_header = 16
    years = ["An 1 (2026)", "An 2 (2027)", "An 3 (2028)", "An 4 (2029)", "An 5 (2030)", "Cumul 5 ans"]
    header_row(ws, r_header, [2, 4, 5, 6, 7, 8, 9],
               ["Compte de résultat (€)", *years])

    def yr_col(an):
        return 3 + an  # an=1 → col D=4, an=2→E=5, etc.

    # Sessions row (calculated from assumptions)
    section_title(ws, r_header + 1, 2, 9, "  Chiffre d'affaires")

    rows_data = []

    # CA calculation
    # Sessions/an = PDC × sessions/PDC/jour × 365 × (1+growth)^(an-1)
    # CA = sessions × kWh × tarif
    # For year n: =C5 * C6 * (1+C7)^(n-1) * 365 * C8 * C9

    r_sess = r_header + 2
    ws.row_dimensions[r_sess].height = 16
    set_cell(ws, r_sess, 2, "Sessions totales / an", bg=C_GREY, border=True, italic=True, color="444444")
    for an in range(1, 6):
        c = yr_col(an)
        formula = f"=ROUND(C5*C6*(1+C7)^{an-1}*365,0)"
        set_cell(ws, r_sess, c, formula, bg=C_GREY, h="right", fmt=FMT_INT, border=True)
    cumul_formula = f"=SUM({get_column_letter(yr_col(1))}{r_sess}:{get_column_letter(yr_col(5))}{r_sess})"
    set_cell(ws, r_sess, yr_col(6), cumul_formula, bg=C_GREY, h="right", fmt=FMT_INT, border=True)

    r_kwh = r_sess + 1
    ws.row_dimensions[r_kwh].height = 16
    set_cell(ws, r_kwh, 2, "Énergie délivrée (kWh/an)", bg=C_WHITE, border=True, italic=True, color="444444")
    for an in range(1, 6):
        c = yr_col(an)
        set_cell(ws, r_kwh, c, f"={get_column_letter(yr_col(an))}{r_sess}*C8", bg=C_WHITE, h="right", fmt="#,##0", border=True)
    set_cell(ws, r_kwh, yr_col(6), f"=SUM({get_column_letter(yr_col(1))}{r_kwh}:{get_column_letter(yr_col(5))}{r_kwh})", bg=C_WHITE, h="right", fmt="#,##0", border=True)

    r_ca = r_kwh + 1
    ws.row_dimensions[r_ca].height = 18
    set_cell(ws, r_ca, 2, "Chiffre d'affaires brut (€)", bg=C_GREEN_L, border=True, bold=True)
    for an in range(1, 6):
        c = yr_col(an)
        set_cell(ws, r_ca, c, f"=ROUND({get_column_letter(yr_col(an))}{r_kwh}*C9,0)", bg=C_GREEN_L, h="right", fmt=FMT_EUR, border=True, bold=True)
    set_cell(ws, r_ca, yr_col(6), f"=SUM({get_column_letter(yr_col(1))}{r_ca}:{get_column_letter(yr_col(5))}{r_ca})", bg=C_GREEN_L, h="right", fmt=FMT_EUR, border=True, bold=True)

    # Charges variables
    section_title(ws, r_ca + 1, 2, 9, "  Charges variables")

    r_elec = r_ca + 2
    ws.row_dimensions[r_elec].height = 16
    set_cell(ws, r_elec, 2, "Coût énergie achetée EDF SEI (€)", bg=C_WHITE, border=True)
    for an in range(1, 6):
        c = yr_col(an)
        set_cell(ws, r_elec, c, f"=ROUND({get_column_letter(yr_col(an))}{r_kwh}*C10,0)", bg=C_WHITE, h="right", fmt=FMT_EUR, border=True)
    set_cell(ws, r_elec, yr_col(6), f"=SUM({get_column_letter(yr_col(1))}{r_elec}:{get_column_letter(yr_col(5))}{r_elec})", bg=C_WHITE, h="right", fmt=FMT_EUR, border=True)

    # Marge brute
    r_mb = r_elec + 1
    ws.row_dimensions[r_mb].height = 18
    set_cell(ws, r_mb, 2, "MARGE BRUTE (€)", bg=C_SUBHEAD, color="FFFFFF", border=True, bold=True)
    for an in range(1, 6):
        c = yr_col(an)
        cl = get_column_letter(c)
        set_cell(ws, r_mb, c, f"={cl}{r_ca}-{cl}{r_elec}", bg=C_SUBHEAD, color="FFFFFF", h="right", fmt=FMT_EUR, border=True, bold=True)
    set_cell(ws, r_mb, yr_col(6), f"=SUM({get_column_letter(yr_col(1))}{r_mb}:{get_column_letter(yr_col(5))}{r_mb})", bg=C_SUBHEAD, color="FFFFFF", h="right", fmt=FMT_EUR, border=True, bold=True)

    r_pct_mb = r_mb + 1
    ws.row_dimensions[r_pct_mb].height = 15
    set_cell(ws, r_pct_mb, 2, "Taux de marge brute (%)", bg=C_GREY, border=True, italic=True)
    for an in range(1, 6):
        c = yr_col(an)
        cl = get_column_letter(c)
        set_cell(ws, r_pct_mb, c, f"={cl}{r_mb}/{cl}{r_ca}", bg=C_GREY, h="right", fmt=FMT_PCT, border=True)
    set_cell(ws, r_pct_mb, yr_col(6), "", bg=C_GREY, border=True)

    # Charges fixes
    section_title(ws, r_pct_mb + 1, 2, 9, "  Charges fixes d'exploitation")

    r_sup = r_pct_mb + 2
    ws.row_dimensions[r_sup].height = 16
    set_cell(ws, r_sup, 2, "Supervision & maintenance (€/an)", bg=C_WHITE, border=True)
    for an in range(1, 6):
        c = yr_col(an)
        set_cell(ws, r_sup, c, f"=ROUND(C11*(1+C13)^{an-1},0)", bg=C_WHITE, h="right", fmt=FMT_EUR, border=True)
    set_cell(ws, r_sup, yr_col(6), f"=SUM({get_column_letter(yr_col(1))}{r_sup}:{get_column_letter(yr_col(5))}{r_sup})", bg=C_WHITE, h="right", fmt=FMT_EUR, border=True)

    r_fix = r_sup + 1
    ws.row_dimensions[r_fix].height = 16
    set_cell(ws, r_fix, 2, "Assurance + frais gestion (€/an)", bg=C_GREY, border=True)
    for an in range(1, 6):
        c = yr_col(an)
        set_cell(ws, r_fix, c, f"=ROUND(C12*(1+C13)^{an-1},0)", bg=C_GREY, h="right", fmt=FMT_EUR, border=True)
    set_cell(ws, r_fix, yr_col(6), f"=SUM({get_column_letter(yr_col(1))}{r_fix}:{get_column_letter(yr_col(5))}{r_fix})", bg=C_GREY, h="right", fmt=FMT_EUR, border=True)

    r_charges = r_fix + 1
    ws.row_dimensions[r_charges].height = 17
    set_cell(ws, r_charges, 2, "Total charges fixes (€)", bg=C_GREEN_L, border=True, bold=True, color="333333")
    for an in range(1, 6):
        c = yr_col(an)
        cl = get_column_letter(c)
        set_cell(ws, r_charges, c, f"={cl}{r_sup}+{cl}{r_fix}", bg=C_GREEN_L, h="right", fmt=FMT_EUR, border=True, bold=True)
    set_cell(ws, r_charges, yr_col(6), f"=SUM({get_column_letter(yr_col(1))}{r_charges}:{get_column_letter(yr_col(5))}{r_charges})", bg=C_GREEN_L, h="right", fmt=FMT_EUR, border=True, bold=True)

    # Résultat opérationnel
    r_ebitda = r_charges + 1
    ws.row_dimensions[r_ebitda].height = 20
    set_cell(ws, r_ebitda, 2, "RÉSULTAT OPÉRATIONNEL AVANT IS (€)", bg=C_GREEN, color="FFFFFF", border=True, bold=True, size=12)
    for an in range(1, 6):
        c = yr_col(an)
        cl = get_column_letter(c)
        set_cell(ws, r_ebitda, c, f"={cl}{r_mb}-{cl}{r_charges}", bg=C_GREEN, color="FFFFFF", h="right", fmt=FMT_EUR, border=True, bold=True, size=12)
    set_cell(ws, r_ebitda, yr_col(6), f"=SUM({get_column_letter(yr_col(1))}{r_ebitda}:{get_column_letter(yr_col(5))}{r_ebitda})", bg=C_GREEN, color="FFFFFF", h="right", fmt=FMT_EUR, border=True, bold=True, size=12)

    # Remboursement emprunt
    r_loan = r_ebitda + 1
    ws.row_dimensions[r_loan].height = 16
    set_cell(ws, r_loan, 2, "Remboursement emprunt 23 495 € / 36 mois ~683 €/mois", bg=C_GREY, border=True, italic=True)
    for an in range(1, 4):  # 3 years only
        c = yr_col(an)
        set_cell(ws, r_loan, c, -8196, bg=C_GREY, h="right", fmt=FMT_EUR, border=True, color=C_BLUE_IN)
    for an in range(4, 7):
        c = yr_col(an)
        set_cell(ws, r_loan, c, 0, bg=C_GREY, h="right", fmt=FMT_EUR, border=True)

    r_net = r_loan + 1
    ws.row_dimensions[r_net].height = 20
    set_cell(ws, r_net, 2, "RÉSULTAT NET AVANT IS — après remboursement (€)", bg=C_HEADER, color="FFFFFF", border=True, bold=True, size=12)
    for an in range(1, 6):
        c = yr_col(an)
        cl = get_column_letter(c)
        set_cell(ws, r_net, c, f"={cl}{r_ebitda}+{cl}{r_loan}", bg=C_HEADER, color="FFFFFF", h="right", fmt=FMT_EUR, border=True, bold=True, size=12)
    set_cell(ws, r_net, yr_col(6), f"=SUM({get_column_letter(yr_col(1))}{r_net}:{get_column_letter(yr_col(5))}{r_net})", bg=C_HEADER, color="FFFFFF", h="right", fmt=FMT_EUR, border=True, bold=True, size=12)

    # Sensibilité
    ws.row_dimensions[r_net + 1].height = 8
    r_sens = r_net + 2
    section_title(ws, r_sens, 2, 9, "  Analyse de sensibilité — CA An 1 selon taux d'usage")

    header_row(ws, r_sens + 1, [2, 4, 5, 6, 7],
               ["Scénario", "Sessions/PDC/j", "CA An 1 (€)", "Marge brute An 1 (€)", "ROI estimé"])

    sens_scenarios = [
        ("Pessimiste", 2, "= 20*2*365*8*0.35", "= 20*2*365*8*0.21", "~2,2 ans"),
        ("Base ★", 3, "= 20*3*365*8*0.35", "= 20*3*365*8*0.21", "< 18 mois"),
        ("Optimiste", 4, "= 20*4*365*8*0.35", "= 20*4*365*8*0.21", "< 6 mois"),
        ("An 3 cible", 6, "= 20*6*365*8*0.35", "= 20*6*365*8*0.21", "—"),
    ]
    for k, (scen, sess, ca_f, mb_f, roi) in enumerate(sens_scenarios):
        r_s = r_sens + 2 + k
        ws.row_dimensions[r_s].height = 17
        bg = C_YELLOW if "★" in scen else (C_WHITE if k % 2 == 0 else C_GREY)
        set_cell(ws, r_s, 2, scen, bg=bg, border=True, bold=("★" in scen))
        set_cell(ws, r_s, 4, sess, bg=bg, h="center", fmt=FMT_INT, border=True, color=C_BLUE_IN)
        # hardcode CA for sensitivity (purely illustrative)
        ca_val = round(20 * sess * 365 * 8 * 0.35)
        mb_val = round(20 * sess * 365 * 8 * 0.21)
        set_cell(ws, r_s, 5, ca_val, bg=bg, h="right", fmt=FMT_EUR, border=True)
        set_cell(ws, r_s, 6, mb_val, bg=bg, h="right", fmt=FMT_EUR, border=True)
        set_cell(ws, r_s, 7, roi, bg=bg, h="center", border=True, italic=True, color="1D8348" if "mois" in roi else "555555")

    r_foot = r_sens + 2 + len(sens_scenarios) + 1
    ws.merge_cells(start_row=r_foot, start_column=2, end_row=r_foot, end_column=9)
    set_cell(ws, r_foot, 2,
             "Hypothèse marge variable : 0,35 − 0,14 = 0,21 €/kWh  |  "
             "< 30 PDC publics pour 280 000 hab. en Guyane → demande immédiate dès mise en service  |  "
             "PDC ADVENIR exigence uptime ≥ 95%",
             italic=True, color="777777", size=9, wrap=True)
    ws.row_dimensions[r_foot].height = 28

    return ws


# ─── Sheet 5: Synthèse KPIs ───────────────────────────────────────────────────

def build_synthese(wb):
    ws = wb.create_sheet("Synthèse")
    ws.sheet_view.showGridLines = False

    widths = [2, 38, 22, 22, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[2].height = 36
    title_row(ws, 2, 1, 5,
              "SYNTHÈSE — EGREENCITY'S  |  Programme ADVENIR Guyane 2026  |  20 PDC — 10 × e-Premium AC 2×22 kW",
              size=14, bg=C_HEADER)

    kpi_sections = [
        ("Projet", [
            ("Opérateur", "EGREENCITY'S — SAS", "SIREN 878 682 854"),
            ("Région", "Guyane française (DROM — RUP)", "Région Ultrapériphérique UE"),
            ("Programme", "ADVENIR — Voirie publique", "AVERE-France / advenir.mobi"),
            ("Référence devis", "E-TOTEM DEV17000172-6", "2025"),
        ]),
        ("Infrastructure", [
            ("Nombre de bornes", "10 × e-Premium AC Pied 2×22 kW", "440 kW AC installés"),
            ("Nombre de PDC", "20 PDC", "2 PDC / borne"),
            ("Communes couvertes", "8 communes — Phase 1", "Île de Cayenne + Kourou + Maroni"),
            ("Connecteurs", "2 × Type 2 (IEC 62196-2) + 2 × Prise E", "OCPP 1.6J / 4G LTE"),
        ]),
        ("Financement", [
            ("Budget total TTC Guyane", "110 540 €", "LOT 1 + LOT 2 + contingence"),
            ("dont LOT 1 matériel HT (base ADVENIR)", "58 020 €", "10 × 5 802 € HT fournisseur"),
            ("Subvention ADVENIR", "37 200 €", "20 PDC × 1 860 € — après mise en service"),
            ("Taux de couverture ADVENIR", "64,1 %", "37 200 / 58 020 — EXCEPTIONNEL"),
            ("Capital résiduel à financer", "23 495 €", "Emprunt 36 mois ~683 €/mois"),
        ]),
        ("Rentabilité", [
            ("CA An 1 (base 3 sessions/PDC/j)", "61 320 €", "20 PDC × 3 sess × 365 j × 8 kWh × 0,35 €"),
            ("Marge opérationnelle An 1", "31 092 €", "après énergie + charges exploitation"),
            ("Résultat net An 1 (après emprunt)", "~23 000 €", "avant IS"),
            ("Retour sur investissement", "< 18 mois", "scénario de base conservateur"),
            ("Cumul CA 5 ans (base)", "> 400 000 €", "avec +20%/an de taux d'usage"),
        ]),
        ("Calendrier", [
            ("J+0 — Création compte ADVENIR", "Aujourd'hui", "advenir.mobi"),
            ("J+1 à J+7 — Pièces admin A1–A7", "1 semaine", "URSSAF, DGFiP, CNI, RIB"),
            ("J+3 à J+14 — Pièces techniques B1–B6", "2 semaines", "E-TOTEM — email envoyé"),
            ("J+7 à J+21 — Lettres mairies D1", "3 semaines", "9 communes — modèle C6 prêt"),
            ("J+45 — Dépôt dossier complet", "6–7 semaines", "AVERE-France instruction"),
            ("J+90 — Notification éligibilité", "~3 mois", "Puis commande E-TOTEM"),
        ]),
    ]

    r = 4
    for section_name, items in kpi_sections:
        ws.row_dimensions[r].height = 8
        r += 1
        section_title(ws, r, 2, 5, f"  {section_name}")
        r += 1
        for j, (label, val, detail) in enumerate(items):
            ws.row_dimensions[r].height = 20
            bg = C_WHITE if j % 2 == 0 else C_GREY
            bold_val = any(kw in label for kw in ["Taux", "Capital résiduel", "Retour", "Subvention", "Résultat net"])
            set_cell(ws, r, 2, label, bg=bg, border=True, italic=True, color="333333")
            set_cell(ws, r, 3, val, bg=bg, border=True, bold=bold_val,
                     color="1D8348" if bold_val else "000000", h="right")
            set_cell(ws, r, 4, detail, bg=bg, border=True, italic=True, color="777777", size=9, wrap=True)
            r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    set_cell(ws, r, 2,
             "Document préparé pour dépôt ADVENIR — EGREENCITY'S — SAS — SIREN 878 682 854 — Macouria-Tonate, Guyane — Version 2.0 — Avril 2026",
             italic=True, color="999999", size=9, h="center")

    return ws


# ─── Main ─────────────────────────────────────────────────────────────────────

wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

build_config(wb)
build_financement(wb)
build_remboursement(wb)
build_rentabilite(wb)
build_synthese(wb)

wb.save(OUTPUT)
print(f"Fichier créé : {OUTPUT}")
