"""
Met à jour _dossiers/02_Dossier_ADVENIR/02_Tableau_20_PDC.xlsx
avec les NOUVEAUX prix 2026 (devis E-TOTEM DEV26000037 du 30/04/2026).

Changement clé :
  Prix HT borne e-Premium AC 2x22 kW : 5 802 € -> 4 948 € HT
  (-854 € HT / borne, soit -8 540 € HT pour 10 bornes)
  Couverture ADVENIR : 64,1% -> 75,2% (sur bornes nues)

Le total ADVENIR (37 200 €) reste inchangé car par PDC.
"""
from openpyxl import load_workbook
from pathlib import Path

XLSX = Path(r"C:\projet\Egreencity\_dossiers\02_Dossier_ADVENIR\02_Tableau_20_PDC.xlsx")

NEW_PRICE_HT = 4948  # nouveau prix unitaire e-Premium AC 2x22 kW (devis DEV26000037)

wb = load_workbook(XLSX)
ws = wb.active

# --- 1. Ligne 3 (indicateur du haut) — mettre à jour le total
# "SIREN ... | 10 stations | 20 PDC | 440 kW | Prime totale : 37 200 EUR"
# Reste inchangée (les bornes sont toujours 10, 20 PDC, 440 kW, Prime 37 200)

# --- 2. Lignes 5 à 24 (données PDC) — colonne N (Prix HT borne) en lignes paires
#       Garder le prix sur le PDC #1 de chaque station, vide sur le PDC #2.
#       Stations sont aux lignes 5, 7, 9, 11, 13, 15, 17, 19, 21, 23 (PDC#1)
for row in range(5, 25):
    cell = ws.cell(row=row, column=14)  # colonne N = Prix HT borne
    if cell.value is not None and isinstance(cell.value, (int, float)):
        cell.value = NEW_PRICE_HT

# --- 3. Ligne 24 (totaux) — la formule SUMIF est dynamique, pas besoin de modifier
#       Mais le texte "= 64,1 %" en colonne O doit etre mis a jour
total_advenir = 37200
total_ht = NEW_PRICE_HT * 10
new_pct = total_advenir / total_ht * 100  # 75.2%
ws.cell(row=24, column=15).value = f"= {new_pct:.1f} % du HT materiel"

# --- 4. Ligne 1 (titre) — pas de changement
# --- 5. Ligne 3 (indicateur synthese du haut) — texte non modifie

# --- 6. Synthese par commune (lignes 30-36) — mettre a jour Prix HT (colonne G)
#       Colonne G = Prix HT, colonne F = Prime ADVENIR
for row in range(30, 37):
    cell_ht = ws.cell(row=row, column=7)  # colonne G
    if cell_ht.value is not None and isinstance(cell_ht.value, (int, float)):
        # Cayenne et Iracoubo ont 2 stations : 2 * 4948 = 9896
        if cell_ht.value == 11604:  # ancienne valeur 2 * 5802
            cell_ht.value = 2 * NEW_PRICE_HT
        elif cell_ht.value == 5802:  # ancienne valeur 1 * 5802
            cell_ht.value = NEW_PRICE_HT

# --- 7. Notes (lignes 41-47) — mettre a jour les references devis
for row in range(40, 50):
    for col in range(1, 16):
        cell = ws.cell(row=row, column=col)
        if cell.value and isinstance(cell.value, str):
            v = cell.value
            v = v.replace("5 802 EUR HT/borne", f"{NEW_PRICE_HT} EUR HT/borne")
            v = v.replace("5802 EUR", f"{NEW_PRICE_HT} EUR")
            v = v.replace("DEV17000172-6", "DEV26000037 du 30/04/2026")
            v = v.replace("64,1%", f"{new_pct:.1f}%")
            v = v.replace("64,1 %", f"{new_pct:.1f} %")
            v = v.replace("58 020 EUR HT", f"{total_ht:,} EUR HT".replace(",", " "))
            v = v.replace("58020 EUR HT", f"{total_ht} EUR HT")
            cell.value = v

# --- 8. Mettre aussi a jour la ligne 3 du haut (indicateur global)
c3 = ws.cell(row=3, column=1)
if c3.value and isinstance(c3.value, str):
    # ne pas changer le contenu, juste vérifier qu'il reste cohérent
    print("Ligne 3:", c3.value)

# Sauvegarde
wb.save(XLSX)
print(f"OK -> {XLSX}")
print(f"Nouveau prix unitaire HT/borne : {NEW_PRICE_HT} EUR")
print(f"Investissement HT total : {total_ht} EUR (10 bornes)")
print(f"ADVENIR : {total_advenir} EUR (inchange)")
print(f"Couverture ADVENIR : {new_pct:.1f} % (vs 64,1 % avant)")
print(f"Reste a charge : {total_ht - total_advenir} EUR HT")
